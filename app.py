
from __future__ import annotations

import csv
from datetime import date, datetime, timedelta
import io
import json
import math
import os
from pathlib import Path
import re
from threading import Lock
import time
from typing import Any
from urllib.parse import quote
from urllib.parse import urlparse
import uuid

from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    abort,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from werkzeug.exceptions import HTTPException

from config import Config
from delivery_preparation_settings import (
    get_delivery_preparation_settings,
    replace_delivery_preparation_settings,
)
from extraction_branches import BRANCHES
from normalize import refresh_missing_warnings
import xml_exporter
import order_store
from auth import (
    authenticate_user,
    create_session,
    get_session_user,
    hash_password,
    revoke_session,
    seed_admin_user,
    session_cookie_name,
    session_cookie_options,
)
from db import _drop_thread_connection, execute, fetch_all, fetch_one, init_db, transaction

load_dotenv()
config = Config.from_env()
OUTPUT_DIR = config.output_dir

app = Flask(__name__)


@app.teardown_appcontext
def _close_db_connection(exc):
    _drop_thread_connection()


_SAFE_ID_RE = re.compile(r"^[A-Za-z0-9._-]+$")
REPLY_EMAIL_TO = (os.getenv("REPLY_EMAIL_TO") or "").strip() or "00primex.eu@gmail.com"
REPLY_EMAIL_BODY = (
    (os.getenv("REPLY_EMAIL_BODY") or "").strip()
    or "Please send the order with furnplan or make the order with 2 positions."
)

_RAW_ALLOWED_ORIGINS = os.getenv("DASHBOARD_ALLOWED_ORIGINS", "")
DASHBOARD_ALLOWED_ORIGINS = {
    origin.strip() for origin in _RAW_ALLOWED_ORIGINS.split(",") if origin.strip()
}
ALLOW_ANY_ORIGIN = "*" in DASHBOARD_ALLOWED_ORIGINS
API_INDEX_CACHE_TTL_SECONDS = max(
    0.5,
    float((os.getenv("API_INDEX_CACHE_TTL_SECONDS") or "3").strip()),
)

EDITABLE_HEADER_FIELDS = [
    "ticket_number",
    "kundennummer",
    "adressnummer",
    "tour",
    "kom_nr",
    "kom_name",
    "liefertermin",
    "wunschtermin",
    "bestelldatum",
    "lieferanschrift",
    "store_name",
    "store_address",
    "seller",
    "delivery_week",
    "iln",
]
EDITABLE_ITEM_FIELDS = ["artikelnummer", "modellnummer", "menge", "furncloud_id"]
HIDDEN_HEADER_EXPORT_FIELDS = {
    "seller",
    "iln",
    "human_review_needed",
    "iln_anl",
    "iln_fil",
    "post_case",
    "reply_needed",
    "adressnummer",
}

VALID_STATUSES = {
    "ok", "human_in_the_loop", "post", "failed", "unknown",
    "waiting_for_reply", "client_replied", "updated_after_reply",
}
VALID_VALIDATION_STATUSES = set(order_store.VALID_VALIDATION_STATUSES)
ALLOWED_SORTS = {"received_at_desc", "received_at_asc"}
ALLOWED_DOWNLOAD_EXTENSIONS = {".xml"}
ALLOWED_DATA_EXPORT_TABLES = frozenset(
    {
        "filialen_import_stage",
        "kunden_import_stage",
        "modelnr_std_import_stage",
        "wochen_import_stage",
    }
)
ALLOWED_DATA_IMPORT_TABLES = frozenset({
    "filialen_import_stage",
    "kunden_import_stage",
    "modelnr_std_import_stage",
})
IMPORT_COLUMN_MAP: dict[str, dict[str, str]] = {
    "kunden_import_stage": {
        "Kundennummer": "kundennummer",
        "Kundenbetrieb": "kundenbetrieb",
        "Name1": "name1",
        "Name2": "name2",
        "Name3": "name3",
        "Strasse": "strasse",
        "Ort": "ort",
        "Postleitzahl": "postleitzahl",
        "Adressnummer": "adressnummer",
        "Tour": "tour",
        "Verband": "verband",
    },
    "filialen_import_stage": {
        "Filial-/Lagerkürzel": "filial_lagerkuerzel",
        "ILN": "iln",
        "Schiene": "schiene",
        "Filiale/Lager": "filiale_lager",
        "Straße": "strasse",
        "PLZ": "plz",
        "Ort": "ort",
        "Rechnungsregulierer (NAD+IV)": "rechnungsregulierer",
        "Gesellschaft": "gesellschaft",
        "Datenempfänger": "datenempfaenger",
    },
    "modelnr_std_import_stage": {
        "VABTRA": "vabtra",
        "VAMDNR": "vamdnr",
    },
}
DATA_EXPORT_TABLE_ALIASES = {
    "fillalen_import_stage": "filialen_import_stage",
}
UNKNOWN_EXTRACTION_BRANCH = "unknown"
KNOWN_EXTRACTION_BRANCH_IDS = frozenset(BRANCHES.keys())
ALLOWED_CLIENT_FILTER_IDS = KNOWN_EXTRACTION_BRANCH_IDS | {UNKNOWN_EXTRACTION_BRANCH}

_ORDER_INDEX_CACHE: dict[str, Any] = {
    "checked_at": 0.0,
    "orders": [],
}
_ORDER_INDEX_LOCK = Lock()
ORDERS_FILTER_COUNTS_CACHE_TTL_SECONDS = max(
    0.5,
    float((os.getenv("ORDERS_FILTER_COUNTS_CACHE_TTL_SECONDS") or "10").strip()),
)
_ORDERS_FILTER_COUNTS_CACHE: dict[tuple[Any, ...], dict[str, Any]] = {}
_ORDERS_FILTER_COUNTS_LOCK = Lock()

init_db()
if seed_admin_user():
    print("Admin user created from environment configuration.")


def _safe_id(value: str) -> str | None:
    if not value or not _SAFE_ID_RE.match(value):
        return None
    return value


def _entry_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return {
            "value": value.get("value", ""),
            "source": value.get("source", ""),
            "confidence": value.get("confidence", ""),
            "derived_from": value.get("derived_from", ""),
        }
    return {"value": value or "", "source": "", "confidence": "", "derived_from": ""}


def _header_value(header: dict[str, Any], key: str) -> str:
    entry = header.get(key, {})
    if isinstance(entry, dict):
        return str(entry.get("value", "") or "")
    return str(entry or "")


def _is_truthy_flag(value: Any) -> bool:
    if isinstance(value, dict):
        value = value.get("value")
    if value is True:
        return True
    return str(value).lower() == "true"


def _reply_mailto(message_id: str, order_id: str, reply_case: str = "") -> str:
    subject = f"Reply needed for order {message_id or order_id}"
    reply_case_section = f"Reply case: {reply_case}\n\n" if reply_case else ""
    body = (
        f"{REPLY_EMAIL_BODY}\n\n"
        f"{reply_case_section}"
        f"Order ID: {order_id}\n"
        f"Message ID: {message_id or order_id}"
    )
    return (
        f"mailto:{REPLY_EMAIL_TO}"
        f"?subject={quote(subject)}"
        f"&body={quote(body)}"
    )


def _reply_case_from_warnings(warnings: list[str]) -> str:
    if not isinstance(warnings, list):
        return ""
    prefix = "Reply needed:"
    for warning in warnings:
        if isinstance(warning, str) and warning.startswith(prefix):
            return warning[len(prefix):].strip()
    return ""


def _manual_entry(value: str) -> dict[str, Any]:
    return {
        "value": value,
        "source": "manual",
        "confidence": 1.0 if value else 0.0,
        "derived_from": "manual_edit",
    }


def _set_manual_entry(target: dict[str, Any], field: str, value: str) -> None:
    entry = target.get(field)
    if not isinstance(entry, dict):
        target[field] = _manual_entry(value)
        return
    entry["value"] = value
    entry["source"] = "manual"
    entry["confidence"] = 1.0 if value else 0.0
    entry["derived_from"] = "manual_edit"


def _clean_form_value(value: str | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_status(value: Any) -> str:
    status = str(value or "ok").strip().lower()
    if status == "partial" or status == "reply":
        return "waiting_for_reply"
    if status not in VALID_STATUSES:
        return "ok"
    return status


def _status_label(value: Any) -> str:
    status = _normalize_status(value)
    labels = {
        "ok": "OK",
        "waiting_for_reply": "Waiting for Reply",
        "human_in_the_loop": "Human in the Loop",
        "post": "Post",
        "failed": "Failed",
        "unknown": "Unknown Client",
    }
    return labels.get(status, "OK")


def _normalize_extraction_branch(value: Any) -> str:
    branch_id = str(value or "").strip().lower()
    if branch_id in ALLOWED_CLIENT_FILTER_IDS:
        return branch_id
    return UNKNOWN_EXTRACTION_BRANCH


def _normalize_client_branches(values: Any) -> set[str]:
    if values is None:
        return set()
    if isinstance(values, str):
        raw_items = [item.strip() for item in values.split(",")]
    elif isinstance(values, (list, tuple, set)):
        raw_items = [str(item).strip() for item in values]
    else:
        return set()
    return {_normalize_extraction_branch(item) for item in raw_items if item}


def _parse_client_branches_input(values: Any) -> tuple[set[str] | None, Any]:
    if values is None:
        return None, None
    if not isinstance(values, (list, tuple, set)):
        return None, _api_error(400, "bad_request", "client_branches must be an array of branch ids")
    parsed = _normalize_client_branches(values)
    raw_values = [str(item or "").strip().lower() for item in values if str(item or "").strip()]
    invalid = sorted(branch for branch in raw_values if branch not in ALLOWED_CLIENT_FILTER_IDS)
    if invalid:
        return None, _api_error(400, "invalid_client", f"Invalid client values: {', '.join(invalid)}")
    return parsed, None


def _extract_client_branches_payload(payload: dict[str, Any]) -> tuple[bool, Any]:
    if "client_branches" in payload:
        return True, payload.get("client_branches")
    if "client_branches" in payload:
        return True, payload.get("client_branches")
    return False, None


def _fetch_user_client_branches(user_id: str) -> list[str]:
    row = fetch_one(
        """
        SELECT COALESCE(array_agg(ucs.branch_id ORDER BY ucs.branch_id), ARRAY[]::text[]) AS client_branches
        FROM user_client_scopes ucs
        WHERE ucs.user_id = %s
        """,
        (user_id,),
    ) or {}
    values = row.get("client_branches")
    if not isinstance(values, list):
        return []
    return [branch for branch in sorted({_normalize_extraction_branch(item) for item in values if str(item or "").strip()})]


def _replace_user_client_scopes(user_id: str, branches: set[str], now: datetime | None = None) -> None:
    execute("DELETE FROM user_client_scopes WHERE user_id = %s", (user_id,))
    if not branches:
        return
    execute(
        """
        INSERT INTO user_client_scopes (user_id, branch_id, created_at)
        SELECT %s, branch_id, %s
        FROM unnest(%s::text[]) AS branch_id
        ON CONFLICT (user_id, branch_id) DO NOTHING
        """,
        (user_id, now or datetime.now().astimezone(), sorted(branches)),
    )


def _order_access_scope(
    user: dict[str, Any] | None,
    *,
    include_assignment: bool = False,
) -> dict[str, Any]:
    principal = user or {}
    is_admin = principal.get("role") == "admin"
    user_id = str(principal.get("id") or "").strip()
    if is_admin:
        return {"is_admin": True, "assigned_user_id": None, "allowed_client_branches": None}
    allowed = _normalize_client_branches(principal.get("client_branches"))
    return {
        "is_admin": False,
        "assigned_user_id": (user_id or None) if include_assignment else None,
        "allowed_client_branches": allowed,
    }


def _effective_client_branches(
    requested: set[str] | None,
    allowed: set[str] | None,
) -> set[str] | None:
    if allowed is None:
        return requested
    if requested is None:
        return set(allowed)
    return requested & allowed


def _serialize_user_record(row: dict[str, Any]) -> dict[str, Any]:
    user_id = str(row.get("id") or "")
    role = str(row.get("role") or "user")
    payload = {
        "id": user_id,
        "username": row.get("username"),
        "email": row.get("email"),
        "role": role,
        "is_active": bool(row.get("is_active")),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "last_login_at": row.get("last_login_at"),
        "client_branches": _fetch_user_client_branches(user_id),
    }
    if role == "admin":
        payload["client_branches"] = []
    return payload


def _parse_received_at(value: Any) -> datetime | None:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.astimezone()


def _effective_received_at(order: dict[str, Any]) -> datetime:
    parsed = _parse_received_at(order.get("received_at"))
    if parsed:
        return parsed

    mtime = order.get("mtime")
    if isinstance(mtime, datetime):
        return mtime.astimezone()

    return datetime.fromtimestamp(0).astimezone()

def _status_counts(orders: list[dict[str, Any]]) -> dict[str, int]:
    counts = {
        "ok": 0, "reply": 0, "human_in_the_loop": 0, "post": 0, "failed": 0,
        "waiting_for_reply": 0, "client_replied": 0, "updated_after_reply": 0,
    }
    for order in orders:
        status = _normalize_status(order.get("status"))
        counts[status] = counts.get(status, 0) + 1
    counts["total"] = len(orders)
    return counts


def _rate(count: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round((count / total) * 100, 2)


def _shift_month_start(value: datetime, delta_months: int) -> datetime:
    month_index = (value.year * 12 + (value.month - 1)) + delta_months
    year = month_index // 12
    month = (month_index % 12) + 1
    return datetime(year, month, 1, tzinfo=value.tzinfo)


def _chart_day_bounds(range_start: datetime, range_end: datetime) -> tuple[datetime, datetime]:
    local_tz = range_start.tzinfo
    chart_start = datetime.combine(range_start.date(), datetime.min.time(), tzinfo=local_tz)
    inclusive_end = max(range_start, range_end - timedelta(microseconds=1))
    chart_end = datetime.combine(inclusive_end.date(), datetime.min.time(), tzinfo=local_tz)
    return chart_start, chart_end


def _chart_month_bounds(range_start: datetime, range_end: datetime) -> tuple[datetime, datetime]:
    inclusive_end = max(range_start, range_end - timedelta(microseconds=1))
    chart_start = datetime(range_start.year, range_start.month, 1, tzinfo=range_start.tzinfo)
    chart_end = datetime(inclusive_end.year, inclusive_end.month, 1, tzinfo=range_start.tzinfo)
    return chart_start, chart_end


def _postgres_timezone_name(value: Any) -> str:
    key = getattr(value, "key", None)
    if key:
      return str(key)
    text = str(value or "").strip()
    if text in {"Central Europe Standard Time", "W. Europe Standard Time"}:
      return "Europe/Budapest"
    return text or "UTC"


def _parse_overview_range(now: datetime) -> tuple[dict[str, Any] | None, Response | None]:
    preset = (request.args.get("range") or "today").strip().lower()
    local_tz = now.tzinfo
    today_start = datetime.combine(now.date(), datetime.min.time(), tzinfo=local_tz)
    selected_year = now.year
    chart_preset = preset

    if preset == "today":
        range_start = today_start
        range_end = now
        chart_preset = "week"
    elif preset == "week":
        range_start = today_start - timedelta(days=now.weekday())
        range_end = now
    elif preset == "month":
        range_start = datetime(now.year, now.month, 1, tzinfo=local_tz)
        range_end = now
    elif preset == "custom_month":
        month_token = (request.args.get("month") or "").strip()
        match = re.fullmatch(r"(\d{4})-(\d{2})", month_token)
        if not match:
            return None, _api_error(400, "invalid_range", "Invalid month. Use YYYY-MM.")
        month_year = int(match.group(1))
        month_number = int(match.group(2))
        if month_number < 1 or month_number > 12:
            return None, _api_error(400, "invalid_range", "Invalid month. Use YYYY-MM.")
        range_start = datetime(month_year, month_number, 1, tzinfo=local_tz)
        range_end = _shift_month_start(range_start, 1)
    elif preset == "3m":
        current_month_start = datetime(now.year, now.month, 1, tzinfo=local_tz)
        range_start = _shift_month_start(current_month_start, -2)
        range_end = now
    elif preset == "6m":
        current_month_start = datetime(now.year, now.month, 1, tzinfo=local_tz)
        range_start = _shift_month_start(current_month_start, -5)
        range_end = now
    elif preset == "year":
        year_token = (request.args.get("year") or "").strip()
        if year_token:
            if not re.fullmatch(r"\d{4}", year_token):
                return None, _api_error(400, "invalid_range", "Invalid year. Use YYYY.")
            selected_year = int(year_token)
        range_start = datetime(selected_year, 1, 1, tzinfo=local_tz)
        range_end = datetime(selected_year + 1, 1, 1, tzinfo=local_tz)
    else:
        return None, _api_error(400, "invalid_range", "Invalid range preset.")

    bucket_granularity = "month" if chart_preset in {"3m", "6m", "year"} else "day"
    if bucket_granularity == "month":
        chart_range_start = range_start
        chart_range_end = range_end
        chart_start, chart_end = _chart_month_bounds(chart_range_start, chart_range_end)
    else:
        if chart_preset == "week":
            chart_range_start = today_start - timedelta(days=now.weekday())
            chart_range_end = now
        else:
            chart_range_start = range_start
            chart_range_end = range_end
        chart_start, chart_end = _chart_day_bounds(chart_range_start, chart_range_end)
    return (
        {
            "preset": preset,
            "month": (request.args.get("month") or "").strip() or None,
            "year": selected_year if preset == "year" else None,
            "start": range_start,
            "end": range_end,
            "chart_range_start": chart_range_start,
            "chart_range_end": chart_range_end,
            "chart_start": chart_start,
            "chart_end": chart_end,
            "bucket_granularity": bucket_granularity,
        },
        None,
    )


def _overview_status_summary_from_counts(
    *,
    total: int,
    ok: int,
    waiting_for_reply: int,
    human_in_the_loop: int,
    post: int,
    unknown: int,
    failed: int,
    updated_after_reply: int,
) -> dict[str, Any]:
    return {
        "total": int(total),
        "statuses": {
            "ok": {"count": int(ok), "rate": _rate(int(ok), int(total))},
            "waiting_for_reply": {
                "count": int(waiting_for_reply),
                "rate": _rate(int(waiting_for_reply), int(total)),
            },
            "human_in_the_loop": {
                "count": int(human_in_the_loop),
                "rate": _rate(int(human_in_the_loop), int(total)),
            },
            "post": {"count": int(post), "rate": _rate(int(post), int(total))},
            "unknown": {"count": int(unknown), "rate": _rate(int(unknown), int(total))},
            "failed": {"count": int(failed), "rate": _rate(int(failed), int(total))},
            "updated_after_reply": {
                "count": int(updated_after_reply),
                "rate": _rate(int(updated_after_reply), int(total)),
            },
        },
    }


def _status_breakdown(orders: list[dict[str, Any]]) -> dict[str, Any]:
    counts = _status_counts(orders)
    total = counts["total"]
    return {
        "total": total,
        "ok": counts["ok"],
        "reply": counts["reply"],
        "human_in_the_loop": counts["human_in_the_loop"],
        "post": counts["post"],
        "failed": counts["failed"],
        "ok_rate": _rate(counts["ok"], total),
        "reply_rate": _rate(counts["reply"], total),
        "human_in_the_loop_rate": _rate(counts["human_in_the_loop"], total),
        "post_rate": _rate(counts["post"], total),
        "failed_rate": _rate(counts["failed"], total),
    }


def _status_breakdown_from_counts(
    *,
    total: int,
    ok: int,
    reply: int,
    human_in_the_loop: int,
    post: int,
    failed: int,
) -> dict[str, Any]:
    return {
        "total": int(total),
        "ok": int(ok),
        "reply": int(reply),
        "human_in_the_loop": int(human_in_the_loop),
        "post": int(post),
        "failed": int(failed),
        "ok_rate": _rate(int(ok), int(total)),
        "reply_rate": _rate(int(reply), int(total)),
        "human_in_the_loop_rate": _rate(int(human_in_the_loop), int(total)),
        "post_rate": _rate(int(post), int(total)),
        "failed_rate": _rate(int(failed), int(total)),
    }


def _invalidate_order_index_cache() -> None:
    with _ORDER_INDEX_LOCK:
        _ORDER_INDEX_CACHE["checked_at"] = 0.0
    with _ORDERS_FILTER_COUNTS_LOCK:
        _ORDERS_FILTER_COUNTS_CACHE.clear()


def _orders_counts_cache_key(
    *,
    q: str,
    received_from: datetime | None,
    received_to: datetime | None,
    statuses: set[str] | None,
    reply_needed: bool | None,
    human_review_needed: bool | None,
    post_case: bool | None,
    validation_statuses: set[str] | None,
    client_branches: set[str] | None,
    delivery_week: str | None,
    assigned_user_id: str | None,
    allowed_client_branches: set[str] | None,
    today_start: datetime,
) -> tuple[Any, ...]:
    return (
        q,
        received_from.isoformat() if received_from else "",
        received_to.isoformat() if received_to else "",
        tuple(sorted(statuses or set())),
        reply_needed,
        human_review_needed,
        post_case,
        tuple(sorted(validation_statuses or set())),
        tuple(sorted(client_branches or set())),
        str(delivery_week or "").strip(),
        assigned_user_id or "",
        tuple(sorted(allowed_client_branches or set())),
        today_start.date().isoformat(),
    )


def _get_cached_orders_counts(cache_key: tuple[Any, ...]) -> dict[str, int] | None:
    now = time.time()
    with _ORDERS_FILTER_COUNTS_LOCK:
        entry = _ORDERS_FILTER_COUNTS_CACHE.get(cache_key)
        if not entry:
            return None
        if now >= float(entry.get("expires_at") or 0):
            _ORDERS_FILTER_COUNTS_CACHE.pop(cache_key, None)
            return None
        snapshot = entry.get("snapshot")
        return dict(snapshot) if isinstance(snapshot, dict) else None


def _store_cached_orders_counts(cache_key: tuple[Any, ...], snapshot: dict[str, int]) -> None:
    now = time.time()
    with _ORDERS_FILTER_COUNTS_LOCK:
        _ORDERS_FILTER_COUNTS_CACHE[cache_key] = {
            "expires_at": now + ORDERS_FILTER_COUNTS_CACHE_TTL_SECONDS,
            "snapshot": dict(snapshot),
        }
        # Keep cache bounded in case many unique filters are requested.
        if len(_ORDERS_FILTER_COUNTS_CACHE) > 128:
            oldest_key = min(
                _ORDERS_FILTER_COUNTS_CACHE.items(),
                key=lambda item: float(item[1].get("expires_at") or 0),
            )[0]
            _ORDERS_FILTER_COUNTS_CACHE.pop(oldest_key, None)


def _get_order_index() -> list[dict[str, Any]]:
    now = time.time()
    with _ORDER_INDEX_LOCK:
        if now - float(_ORDER_INDEX_CACHE["checked_at"]) < API_INDEX_CACHE_TTL_SECONDS:
            return list(_ORDER_INDEX_CACHE["orders"])

        try:
            orders = order_store.list_order_summaries()
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"Failed to load order summaries from DB: {exc}") from exc
        _ORDER_INDEX_CACHE["checked_at"] = now
        _ORDER_INDEX_CACHE["orders"] = orders
        return list(orders)


def _serialize_order_summary(order: dict[str, Any]) -> dict[str, Any]:
    effective_received_at = _effective_received_at(order)
    mtime = order.get("mtime")
    mtime_text = mtime.isoformat() if isinstance(mtime, datetime) else ""

    return {
        "id": order.get("id", ""),
        "file_name": order.get("file_name", ""),
        "message_id": order.get("message_id", ""),
        "received_at": order.get("received_at", ""),
        "effective_received_at": effective_received_at.isoformat(),
        "status": _normalize_status(order.get("status")),
        "item_count": int(order.get("item_count") or 0),
        "warnings_count": int(order.get("warnings_count") or 0),
        "errors_count": int(order.get("errors_count") or 0),
        "ticket_number": order.get("ticket_number", ""),
        "kundennummer": order.get("kundennummer", ""),
        "kom_nr": order.get("kom_nr", ""),
        "kom_name": order.get("kom_name", ""),
        "liefertermin": order.get("liefertermin", ""),
        "wunschtermin": order.get("wunschtermin", ""),
        "delivery_week": order.get("delivery_week", ""),
        "store_name": order.get("store_name", ""),
        "store_address": order.get("store_address", ""),
        "iln": order.get("iln", ""),
        "extraction_branch": _normalize_extraction_branch(order.get("extraction_branch")),
        "human_review_needed": bool(order.get("human_review_needed")),
        "reply_needed": bool(order.get("reply_needed")),
        "post_case": bool(order.get("post_case")),
        "validation_status": order_store.normalize_validation_status(order.get("validation_status")),
        "validation_summary": order.get("validation_summary", ""),
        "validation_checked_at": order.get("validation_checked_at", ""),
        "validation_provider": order.get("validation_provider", ""),
        "validation_model": order.get("validation_model", ""),
        "validation_stale_reason": order.get("validation_stale_reason", ""),
        "reply_mailto": order.get("reply_mailto", ""),
        "parse_error": order.get("parse_error"),
        "mtime": mtime_text,
    }

def _parse_bool_query(value: str | None) -> bool | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized == "":
        return None
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    return None


def _parse_date_query(value: str | None) -> date | None:
    if value is None:
        return None
    text = value.strip()
    if not text:
        return None
    return datetime.strptime(text, "%Y-%m-%d").date()


def _api_error(status_code: int, code: str, message: str):
    return jsonify({"error": {"code": code, "message": message}}), status_code


def _order_store_error_response(exc: order_store.OrderStoreError):
    return _api_error(exc.status_code, exc.code, exc.message)


def _mark_order_validation_stale(order_id: str, *, actor_user_id: str | None, reason: str) -> None:
    try:
        order_store.mark_validation_stale(
            order_id=order_id,
            reason=reason,
            actor_user_id=actor_user_id,
        )
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"Failed to mark validation stale: {exc}") from exc


def _response_status_code(error_response: Any, default: int = 500) -> int:
    if isinstance(error_response, tuple) and len(error_response) >= 2:
        code = error_response[1]
        if isinstance(code, int):
            return code
    if isinstance(error_response, Response):
        return int(error_response.status_code or default)
    return default


def _require_admin() -> Any:
    user = getattr(g, "user", None) or {}
    if user.get("role") != "admin":
        return _api_error(403, "forbidden", "Admin access required")
    return None


def require_auth(req) -> Any:
    if req.method == "OPTIONS":
        return None

    session_id = req.cookies.get(session_cookie_name(), "")
    user = get_session_user(session_id)
    if not user:
        return _api_error(401, "unauthorized", "Authentication required")

    g.user = user
    g.session_id = session_id
    return None


def _is_origin_allowed(origin: str | None) -> bool:
    if not origin:
        return False
    if ALLOW_ANY_ORIGIN:
        return True
    parsed_origin = urlparse(origin)
    if parsed_origin.scheme in {"http", "https"} and parsed_origin.hostname in {"localhost", "127.0.0.1"}:
        return True
    return origin in DASHBOARD_ALLOWED_ORIGINS


def _append_vary(existing: str | None, value: str) -> str:
    if not existing:
        return value
    parts = [part.strip() for part in existing.split(",") if part.strip()]
    if value in parts:
        return ", ".join(parts)
    parts.append(value)
    return ", ".join(parts)


def _filter_orders(
    orders: list[dict[str, Any]],
    *,
    q: str,
    date_from: date | None,
    date_to: date | None,
    statuses: set[str] | None,
    reply_needed: bool | None,
    human_review_needed: bool | None,
    post_case: bool | None,
    client_branches: set[str] | None,
) -> list[dict[str, Any]]:
    query = q.strip().lower()
    result: list[dict[str, Any]] = []

    for order in orders:
        effective_dt = _effective_received_at(order)
        effective_date = effective_dt.date()

        if date_from and effective_date < date_from:
            continue
        if date_to and effective_date > date_to:
            continue

        status = _normalize_status(order.get("status"))
        if statuses and status not in statuses:
            continue

        if reply_needed is not None and bool(order.get("reply_needed")) != reply_needed:
            continue
        if human_review_needed is not None and bool(order.get("human_review_needed")) != human_review_needed:
            continue
        if post_case is not None and bool(order.get("post_case")) != post_case:
            continue

        extraction_branch = _normalize_extraction_branch(order.get("extraction_branch"))
        if client_branches and extraction_branch not in client_branches:
            continue

        if query:
            searchable = " ".join(
                [
                    str(order.get("ticket_number") or ""),
                    str(order.get("kom_nr") or ""),
                    str(order.get("kom_name") or ""),
                    str(order.get("message_id") or ""),
                    str(order.get("file_name") or ""),
                ]
            ).lower()
            if query not in searchable:
                continue

        cloned = dict(order)
        cloned["extraction_branch"] = extraction_branch
        cloned["_effective_dt"] = effective_dt
        result.append(cloned)

    return result


def _sort_orders(orders: list[dict[str, Any]], sort_key: str) -> list[dict[str, Any]]:
    reverse = sort_key != "received_at_asc"
    return sorted(
        orders,
        key=lambda order: order.get("_effective_dt") or _effective_received_at(order),
        reverse=reverse,
    )


def _tab_counts(orders: list[dict[str, Any]]) -> dict[str, int]:
    today = datetime.now().astimezone().date()
    return {
        "all": len(orders),
        "today": sum(1 for order in orders if _effective_received_at(order).date() == today),
        "waiting_for_reply": sum(
            1 for order in orders if _normalize_status(order.get("status")) == "waiting_for_reply"
        ),
        "manual_review": sum(
            1 for order in orders if _normalize_status(order.get("status")) == "human_in_the_loop"
        ),
        "unknown": sum(
            1 for order in orders if _normalize_status(order.get("status")) == "unknown"
        ),
        "updated_after_reply": sum(
            1 for order in orders if order.get("status") == "updated_after_reply"
        ),
    }


def _parse_orders_query(allow_default_pagination: bool = True):
    q = (request.args.get("q") or "").strip()
    queue = (request.args.get("queue") or "").strip().lower()
    if queue and queue != "today":
        return None, _api_error(400, "invalid_queue", f"Invalid queue value '{queue}'")

    raw_status = (request.args.get("status") or "").strip()
    statuses: set[str] | None = None
    if raw_status:
        parsed_statuses = {item.strip().lower() for item in raw_status.split(",") if item.strip()}
        invalid = [status for status in parsed_statuses if status not in VALID_STATUSES]
        if invalid:
            return None, _api_error(400, "invalid_status", f"Invalid status values: {', '.join(sorted(invalid))}")
        statuses = {_normalize_status(status) for status in parsed_statuses}

    try:
        date_from = _parse_date_query(request.args.get("from"))
    except ValueError:
        return None, _api_error(400, "invalid_date", "Invalid 'from' date format. Use YYYY-MM-DD.")

    try:
        date_to = _parse_date_query(request.args.get("to"))
    except ValueError:
        return None, _api_error(400, "invalid_date", "Invalid 'to' date format. Use YYYY-MM-DD.")

    if date_from and date_to and date_from > date_to:
        return None, _api_error(400, "invalid_date_range", "'from' cannot be after 'to'.")

    raw_reply_needed = request.args.get("reply_needed")
    reply_needed = _parse_bool_query(raw_reply_needed)
    if raw_reply_needed not in (None, "") and reply_needed is None:
        return None, _api_error(400, "invalid_flag", "Invalid reply_needed flag. Use true or false.")

    raw_human_review = request.args.get("human_review_needed")
    human_review_needed = _parse_bool_query(raw_human_review)
    if raw_human_review not in (None, "") and human_review_needed is None:
        return None, _api_error(400, "invalid_flag", "Invalid human_review_needed flag. Use true or false.")

    raw_post_case = request.args.get("post_case")
    post_case = _parse_bool_query(raw_post_case)
    if raw_post_case not in (None, "") and post_case is None:
        return None, _api_error(400, "invalid_flag", "Invalid post_case flag. Use true or false.")

    raw_validation_status = (request.args.get("validation_status") or "").strip()
    validation_statuses: set[str] | None = None
    if raw_validation_status:
        parsed_validation_statuses = {
            item.strip().lower() for item in raw_validation_status.split(",") if item.strip()
        }
        invalid = [
            status for status in parsed_validation_statuses if status not in VALID_VALIDATION_STATUSES
        ]
        if invalid:
            return None, _api_error(
                400,
                "invalid_validation_status",
                f"Invalid validation_status values: {', '.join(sorted(invalid))}",
            )
        validation_statuses = {
            order_store.normalize_validation_status(status) for status in parsed_validation_statuses
        }

    raw_client = (request.args.get("client") or "").strip()
    client_branches: set[str] | None = None
    if raw_client:
        parsed_client_branches = {item.strip().lower() for item in raw_client.split(",") if item.strip()}
        invalid = sorted(branch for branch in parsed_client_branches if branch not in ALLOWED_CLIENT_FILTER_IDS)
        if invalid:
            return None, _api_error(400, "invalid_client", f"Invalid client values: {', '.join(invalid)}")
        client_branches = parsed_client_branches

    delivery_week = (request.args.get("delivery_week") or "").strip() or None

    sort_key = (request.args.get("sort") or "received_at_desc").strip().lower()
    if sort_key not in ALLOWED_SORTS:
        return None, _api_error(
            400,
            "invalid_sort",
            f"Unsupported sort '{sort_key}'. Allowed: {', '.join(sorted(ALLOWED_SORTS))}.",
        )

    page = 1
    page_size = 100
    should_paginate = allow_default_pagination
    if not allow_default_pagination and "page" not in request.args and "page_size" not in request.args:
        should_paginate = False

    if "page" in request.args:
        try:
            page = max(1, int((request.args.get("page") or "1").strip()))
        except ValueError:
            return None, _api_error(400, "invalid_pagination", "Invalid page value.")

    if "page_size" in request.args:
        try:
            page_size = int((request.args.get("page_size") or "100").strip())
        except ValueError:
            return None, _api_error(400, "invalid_pagination", "Invalid page_size value.")
        if page_size < 1 or page_size > 500:
            return None, _api_error(400, "invalid_pagination", "page_size must be between 1 and 500.")
    elif should_paginate:
        page_size = 100

    query = {
        "q": q,
        "queue": queue or None,
        "date_from": date_from,
        "date_to": date_to,
        "statuses": statuses,
        "reply_needed": reply_needed,
        "human_review_needed": human_review_needed,
        "post_case": post_case,
        "validation_statuses": validation_statuses,
        "client_branches": client_branches,
        "delivery_week": delivery_week,
        "sort_key": sort_key,
        "page": page,
        "page_size": page_size,
        "paginate": should_paginate,
    }
    return query, None


def _query_orders(
    allow_default_pagination: bool = True,
    *,
    access_scope: dict[str, Any] | None = None,
) -> tuple[dict[str, Any] | None, Any]:
    parsed, parse_error = _parse_orders_query(allow_default_pagination=allow_default_pagination)
    if parse_error is not None:
        return None, parse_error

    local_tz = datetime.now().astimezone().tzinfo
    today_local = datetime.now().astimezone().date()
    today_start = datetime.combine(today_local, datetime.min.time(), tzinfo=local_tz)
    today_end = today_start + timedelta(days=1)

    received_from = None
    if parsed["date_from"] is not None:
        received_from = datetime.combine(parsed["date_from"], datetime.min.time(), tzinfo=local_tz)

    received_to = None
    if parsed["date_to"] is not None:
        received_to = datetime.combine(parsed["date_to"] + timedelta(days=1), datetime.min.time(), tzinfo=local_tz)

    counts_received_from = received_from
    counts_received_to = received_to

    if parsed["queue"] == "today" and parsed["date_from"] is None and parsed["date_to"] is None:
        received_from = today_start
        received_to = today_end

    scope = access_scope or {"assigned_user_id": None, "allowed_client_branches": None}
    effective_client_branches = _effective_client_branches(
        parsed["client_branches"],
        scope.get("allowed_client_branches"),
    )

    counts_cache_key = _orders_counts_cache_key(
        q=parsed["q"],
        received_from=received_from,
        received_to=received_to,
        statuses=None,
        reply_needed=parsed["reply_needed"],
        human_review_needed=parsed["human_review_needed"],
        post_case=parsed["post_case"],
        validation_statuses=None,
        client_branches=effective_client_branches,
        delivery_week=parsed["delivery_week"],
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
        today_start=today_start,
    )
    cached_counts = _get_cached_orders_counts(counts_cache_key)

    try:
        query_result = order_store.query_order_summaries(
            q=parsed["q"],
            received_from=received_from,
            received_to=received_to,
            counts_received_from=counts_received_from,
            counts_received_to=counts_received_to,
            statuses=parsed["statuses"],
            reply_needed=parsed["reply_needed"],
            human_review_needed=parsed["human_review_needed"],
            post_case=parsed["post_case"],
            validation_statuses=parsed["validation_statuses"],
            client_branches=effective_client_branches,
            delivery_week=parsed["delivery_week"],
            assigned_user_id=scope.get("assigned_user_id"),
            allowed_client_branches=scope.get("allowed_client_branches"),
            sort_key=parsed["sort_key"],
            page=parsed["page"],
            page_size=parsed["page_size"],
            paginate=parsed["paginate"],
            today_start=today_start,
            today_end=today_end,
            counts_override=cached_counts,
        )
    except Exception as exc:  # noqa: BLE001
        return None, _api_error(500, "db_error", f"Failed to query orders: {exc}")

    count_snapshot = query_result.get("count_snapshot")
    if isinstance(count_snapshot, dict):
        _store_cached_orders_counts(counts_cache_key, count_snapshot)

    rows = query_result["orders"]
    return (
        {
            "orders": rows,
            "orders_serialized": [_serialize_order_summary(order) for order in rows],
            "pagination": query_result["pagination"],
            "counts": query_result["counts"],
        },
        None,
    )

def _header_val(header: dict[str, Any], key: str) -> str:
    entry = header.get(key)
    if isinstance(entry, dict):
        return str(entry.get("value", "") or "").strip()
    return str(entry or "").strip()


def _sanitize_xml_base(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip())
    return cleaned.strip("_") or ""


def _resolve_xml_files(order_id: str, header: dict[str, Any]) -> list[dict[str, str]]:
    effective_base = (
        _sanitize_xml_base(_header_val(header, "ticket_number"))
        or _sanitize_xml_base(_header_val(header, "kom_nr"))
        or _sanitize_xml_base(_header_val(header, "kom_name"))
        or "unknown"
    )
    order_info_xml = f"OrderInfo_{effective_base}.xml"
    article_info_xml = f"OrderArticleInfo_{effective_base}.xml"
    if not (OUTPUT_DIR / order_info_xml).exists():
        order_info_xml = f"OrderInfo_{order_id}.xml"
    if not (OUTPUT_DIR / article_info_xml).exists():
        article_info_xml = f"OrderArticleInfo_{order_id}.xml"

    xml_files: list[dict[str, str]] = []
    if (OUTPUT_DIR / order_info_xml).exists():
        xml_files.append({"name": "Order Info XML", "filename": order_info_xml})
    if (OUTPUT_DIR / article_info_xml).exists():
        xml_files.append({"name": "Article Info XML", "filename": article_info_xml})
    return xml_files


def _load_order(
    order_id: str,
    *,
    assigned_user_id: str | None = None,
    allowed_client_branches: set[str] | None = None,
) -> tuple[dict[str, Any] | None, Any]:
    safe_id = _safe_id(order_id)
    if not safe_id:
        return None, _api_error(404, "not_found", "Order not found")

    try:
        uuid.UUID(safe_id)
    except ValueError:
        return None, _api_error(404, "not_found", "Order not found")
    try:
        order = order_store.get_order_detail(
            safe_id,
            assigned_user_id=assigned_user_id,
            allowed_client_branches=allowed_client_branches,
        )
    except Exception as exc:  # noqa: BLE001
        return None, _api_error(500, "db_error", f"Failed to load order: {exc}")
    if not order:
        return None, _api_error(404, "not_found", "Order not found")
    order["reply_mailto"] = (
        _reply_mailto(order.get("message_id") or safe_id, safe_id, _reply_case_from_warnings(order.get("warnings", [])))
        if order.get("reply_needed")
        else ""
    )
    return order, None


def _order_api_payload(order: dict[str, Any]) -> dict[str, Any]:
    data = order["data"]
    response = dict(data)
    # Override status with authoritative DB column value.
    # The payload-derived status (from derive_status/header flags) can be stale
    # after reply processing — the DB status is always up-to-date.
    response["status"] = _normalize_status(order.get("status"))
    response["extraction_branch"] = _normalize_extraction_branch(response.get("extraction_branch"))
    response["order_id"] = order["safe_id"]
    response["header"] = order["header"]
    response["items"] = order["items"]
    response["warnings"] = order["warnings"]
    response["errors"] = order["errors"]
    response["parse_error"] = order["parse_error"]
    response["xml_files"] = _resolve_xml_files(order["safe_id"], order["header"])
    response["is_editable"] = bool(order.get("is_editable", order["human_review_needed"] and not order["parse_error"]))
    response["editability_reason"] = str(order.get("editability_reason") or "")
    response["reply_mailto"] = order["reply_mailto"]
    response["reply_needed"] = order["reply_needed"]
    response["post_case"] = order["post_case"]
    response["validation_status"] = order_store.normalize_validation_status(order.get("validation_status"))
    response["validation_summary"] = str(order.get("validation_summary") or "")
    response["validation_checked_at"] = str(order.get("validation_checked_at") or "")
    response["validation_provider"] = str(order.get("validation_provider") or "")
    response["validation_model"] = str(order.get("validation_model") or "")
    response["validation_stale_reason"] = str(order.get("validation_stale_reason") or "")
    response["validation_issues"] = order.get("validation_issues") if isinstance(order.get("validation_issues"), list) else []
    response["review_task_id"] = order.get("review_task_id")
    response["review_state"] = order.get("review_state")
    response["assigned_user"] = order.get("assigned_user")
    response["claim_expires_at"] = order.get("claim_expires_at")
    response["sla_due_at"] = order.get("sla_due_at")
    response["last_event_at"] = order.get("last_event_at")
    response["editable_header_fields"] = EDITABLE_HEADER_FIELDS
    response["editable_item_fields"] = EDITABLE_ITEM_FIELDS
    return response


def _export_entry_value(entry: Any) -> Any:
    value = entry.get("value", "") if isinstance(entry, dict) else entry
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _export_entry_confidence(entry: Any) -> float | None:
    if not isinstance(entry, dict):
        return None
    try:
        confidence = float(entry.get("confidence"))
    except (TypeError, ValueError):
        return None
    return confidence if math.isfinite(confidence) else None


def _ensure_string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    if value in (None, ""):
        return []
    return [str(value)]


def _load_order_export_data(
    order: dict[str, Any],
    *,
    payload_map: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    order_id = str(order.get("id") or "")
    if not order_id:
        raise ValueError("Missing order ID for export payload")

    file_name = str(order.get("file_name") or "")
    detail = (payload_map or {}).get(order_id)
    if detail is None:
        detail = order_store.get_order_detail(order_id)
    if not detail:
        raise ValueError(f"Order not found for export payload: {order_id}")

    payload = detail.get("data")
    data = payload if isinstance(payload, dict) else {}
    parse_error = str(detail.get("parse_error") or "")

    header = data.get("header") if isinstance(data.get("header"), dict) else {}
    items = data.get("items") if isinstance(data.get("items"), list) else []
    warnings = _ensure_string_list(data.get("warnings", order.get("warnings", [])))
    errors = _ensure_string_list(data.get("errors", order.get("errors", [])))

    return {
        "order_id": order_id,
        "file_name": file_name,
        "message_id": str(data.get("message_id") or order.get("message_id") or order_id),
        "received_at": str(data.get("received_at") or order.get("received_at") or ""),
        "status": _normalize_status(data.get("status") or order.get("status")),
        "item_count": len(items),
        "warnings_count": len(warnings),
        "errors_count": len(errors),
        "reply_needed": _is_truthy_flag(header.get("reply_needed")) if header else bool(order.get("reply_needed")),
        "human_review_needed": _is_truthy_flag(header.get("human_review_needed"))
        if header
        else bool(order.get("human_review_needed")),
        "post_case": _is_truthy_flag(header.get("post_case")) if header else bool(order.get("post_case")),
        "validation_status": str(detail.get("validation_status") or order.get("validation_status") or "not_run"),
        "validation_summary": str(detail.get("validation_summary") or order.get("validation_summary") or ""),
        "validation_checked_at": str(detail.get("validation_checked_at") or order.get("validation_checked_at") or ""),
        "validation_provider": str(detail.get("validation_provider") or order.get("validation_provider") or ""),
        "validation_model": str(detail.get("validation_model") or order.get("validation_model") or ""),
        "validation_stale_reason": str(detail.get("validation_stale_reason") or order.get("validation_stale_reason") or ""),
        "validation_issues": detail.get("validation_issues") if isinstance(detail.get("validation_issues"), list) else [],
        "warnings": warnings,
        "errors": errors,
        "parse_error": parse_error or "",
        "header": header,
        "items": items,
    }


def _preload_order_export_payloads(orders: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    order_ids = [str(order.get("id") or "").strip() for order in orders]
    order_ids = [order_id for order_id in order_ids if order_id]
    if not order_ids:
        return {}
    return order_store.get_order_payload_map(order_ids)


def _as_orders_xlsx_bytes(
    orders: list[dict[str, Any]],
    *,
    title: str = "Orders",
    initials: str = "",
) -> bytes:
    payload_map = _preload_order_export_payloads(orders)
    parsed_orders = [_load_order_export_data(order, payload_map=payload_map) for order in orders]

    def _format_sheet(
        target_sheet,
        *,
        sheet_title: str,
        table_columns: list[str],
        data_rows: list[list[Any]],
        status_column_name: str | None = None,
    ) -> None:
        now = datetime.now().astimezone()
        date_text = now.strftime("%d.%m.%Y %H:%M")
        last_col = len(table_columns)
        last_col_letter = get_column_letter(last_col)

        target_sheet.merge_cells(f"A1:{last_col_letter}1")
        title_cell = target_sheet["A1"]
        title_cell.value = sheet_title.upper()
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        date_cell = target_sheet.cell(row=2, column=last_col, value=date_text)
        date_cell.alignment = Alignment(horizontal="right", vertical="bottom")

        header_row_index = 4
        header_alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        body_alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
        for col_index, column_name in enumerate(table_columns, start=1):
            cell = target_sheet.cell(row=header_row_index, column=col_index, value=str(column_name).upper())
            cell.font = Font(bold=True)
            cell.alignment = header_alignment
            cell.fill = header_fill

        data_start_row = header_row_index + 1
        for row_offset, row_values in enumerate(data_rows, start=0):
            row_number = data_start_row + row_offset
            row_values[0] = row_offset + 1
            for col_index, value in enumerate(row_values, start=1):
                cell = target_sheet.cell(row=row_number, column=col_index, value=value)
                cell.alignment = body_alignment

        if not data_rows:
            data_start_row = header_row_index + 1

        last_row = max(header_row_index, data_start_row + len(data_rows) - 1)
        target_sheet.freeze_panes = "B5"
        target_sheet.auto_filter.ref = f"A{header_row_index}:{last_col_letter}{last_row}"
        target_sheet.print_title_rows = f"{header_row_index}:{header_row_index}"

        thin = Side(style="thin", color="000000")
        thick = Side(style="thick", color="000000")
        thin_border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for col_index in range(1, last_col + 1):
            cell = target_sheet.cell(row=header_row_index, column=col_index)
            border = Border(
                top=thick,
                bottom=thick,
                left=thick if col_index == 1 else None,
                right=thick if col_index == last_col else None,
            )
            cell.border = border

        nr_font = Font(bold=True)
        for row_idx in range(data_start_row, last_row + 1):
            cell = target_sheet.cell(row=row_idx, column=1)
            cell.font = nr_font

        for row_idx in range(header_row_index, last_row + 1):
            for col_index in range(1, last_col + 1):
                cell = target_sheet.cell(row=row_idx, column=col_index)
                if row_idx != header_row_index:
                    cell.border = thin_border

        numeric_format = "#,##0"
        currency_format = "#,##0.00"
        currency_markers = {"amount", "total", "price", "value", "sum"}
        for row_idx in range(data_start_row, last_row + 1):
            for col_index, column_name in enumerate(table_columns, start=1):
                cell = target_sheet.cell(row=row_idx, column=col_index)
                if isinstance(cell.value, (int, float)) and column_name.lower() not in {"nr", "punoi"}:
                    if column_name.lower() in currency_markers or isinstance(cell.value, float):
                        cell.number_format = currency_format
                    else:
                        cell.number_format = numeric_format

        if status_column_name:
            status_idx = None
            for col_index, column_name in enumerate(table_columns, start=1):
                if str(column_name).strip().lower() == status_column_name.strip().lower():
                    status_idx = col_index
                    break
            if status_idx is not None:
                status_fills = {
                    "ok": PatternFill(fill_type="solid", fgColor="C6EFCE"),
                    "reply": PatternFill(fill_type="solid", fgColor="FCE4D6"),
                    "human_in_the_loop": PatternFill(fill_type="solid", fgColor="FFEB9C"),
                    "post": PatternFill(fill_type="solid", fgColor="DDEBF7"),
                    "failed": PatternFill(fill_type="solid", fgColor="F8CBAD"),
                }
                for row_idx in range(data_start_row, last_row + 1):
                    cell = target_sheet.cell(row=row_idx, column=status_idx)
                    status_value = str(cell.value or "").strip().lower()
                    fill = status_fills.get(status_value)
                    if fill is not None:
                        cell.fill = fill

        max_width = 45
        min_width = 8
        padding = 2
        header_widths = {idx: len(str(name)) for idx, name in enumerate(table_columns, start=1)}
        column_widths = dict(header_widths)
        for row_values in data_rows:
            for col_index, value in enumerate(row_values, start=1):
                text = "" if value is None else str(value)
                column_widths[col_index] = max(column_widths[col_index], len(text))
        for col_index, raw_width in column_widths.items():
            width = max(min_width, min(max_width, raw_width + padding))
            width = max(width, header_widths.get(col_index, 0) + padding)
            target_sheet.column_dimensions[get_column_letter(col_index)].width = width

        target_sheet.page_margins = PageMargins(
            top=0.36,
            bottom=0.51,
            left=0.1,
            right=0.1,
            header=0.15,
            footer=0.2,
        )
        target_sheet.page_setup.fitToWidth = 1
        target_sheet.page_setup.fitToHeight = 0
        target_sheet.oddFooter.center.text = "Page &P/&N"
        target_sheet.oddFooter.right.text = ""

    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Orders"

    header_fields = [
        ("ticket_number", "Ticket Number"),
        ("kundennummer", "Kundennummer"),
        ("tour", "Tour"),
        ("kom_nr", "Kom Nr"),
        ("liefertermin", "Liefertermin"),
        ("wunschtermin", "Wunschtermin"),
        ("bestelldatum", "Bestelldatum"),
        ("lieferanschrift", "Lieferanschrift"),
        ("store_name", "Store Name"),
        ("store_address", "Store Address"),
        ("delivery_week", "Delivery Week"),
        ("mail_to", "Mail To"),
    ]
    orders_columns = ["Nr"] + [label for _, label in header_fields] + ["Items"]
    orders_rows: list[list[Any]] = []

    for parsed_order in parsed_orders:
        header = parsed_order["header"]
        items = [item for item in parsed_order.get("items", []) if isinstance(item, dict)]
        item_lines: list[str] = []
        for index, item in enumerate(items, start=1):
            quantity = _export_entry_value(item.get("menge", "")) or "-"
            article_number = _export_entry_value(item.get("artikelnummer", "")) or "-"
            model_number = _export_entry_value(item.get("modellnummer", "")) or "-"
            item_lines.append(f"{index}. {quantity} - {article_number} {model_number}".strip())

        orders_rows.append(
            [None]
            + [(_export_entry_value(header.get(field_key, "")) or "-") for field_key, _ in header_fields]
            + ["\n".join(item_lines) if item_lines else "-"]
        )

    title_text = (title or "Orders").strip() or "Orders"
    _format_sheet(
        orders_sheet,
        sheet_title=title_text,
        table_columns=orders_columns,
        data_rows=orders_rows,
    )

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _header_export_label(field: str) -> str:
    return str(field or "").replace("_", " ").title()


def _visible_order_header_rows(order: dict[str, Any]) -> list[tuple[str, Any]]:
    header = order.get("header")
    if not isinstance(header, dict):
        return []

    rows: list[tuple[str, Any]] = []
    seen: set[str] = set()

    for field in EDITABLE_HEADER_FIELDS:
        field_key = str(field or "")
        if field_key.lower() in HIDDEN_HEADER_EXPORT_FIELDS:
            continue
        if field_key in header:
            rows.append((field_key, header[field_key]))
            seen.add(field_key)

    for field_key in sorted(str(key) for key in header.keys() if str(key) not in seen):
        if field_key.lower() in HIDDEN_HEADER_EXPORT_FIELDS:
            continue
        rows.append((field_key, header[field_key]))

    return rows


def _as_order_header_xlsx_bytes(order: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Order Export"

    header = order.get("header") if isinstance(order.get("header"), dict) else {}
    items = order.get("items") if isinstance(order.get("items"), list) else []

    header_fields = [
        ("ticket_number", "Ticket Number"),
        ("kundennummer", "Kundennummer"),
        ("tour", "Tour"),
        ("kom_nr", "Kom Nr"),
        ("liefertermin", "Liefertermin"),
        ("wunschtermin", "Wunschtermin"),
        ("bestelldatum", "Bestelldatum"),
        ("lieferanschrift", "Lieferanschrift"),
        ("store_name", "Store Name"),
        ("store_address", "Store Address"),
        ("delivery_week", "Delivery Week"),
        ("mail_to", "Mail To"),
    ]
    item_fields = [
        ("artikelnummer", "Article Number"),
        ("modellnummer", "Model Number"),
        ("menge", "Quantity"),
        ("furncloud_id", "Furncloud ID"),
    ]
    columns = [label for _, label in header_fields] + [label for _, label in item_fields]
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    body_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="000000")
    thin_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    title = f"HEADER INFORMATION - {order.get('safe_id') or order.get('id') or ''}".strip(" -")
    last_col_letter = get_column_letter(len(columns))
    sheet.merge_cells(f"A1:{last_col_letter}1")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="bottom")

    exported_at = datetime.now().astimezone().strftime("%d.%m.%Y %H:%M")
    sheet.cell(row=2, column=len(columns), value=exported_at).alignment = Alignment(horizontal="right", vertical="bottom")

    header_row_index = 4
    for col_index, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row_index, column=col_index, value=column_name.upper())
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_rows = [item for item in items if isinstance(item, dict)] or [{}]
    for index, item in enumerate(data_rows, start=1):
        row_values = [
            _export_entry_value(header.get(field_key, "")) or "-"
            for field_key, _ in header_fields
        ] + [
            _export_entry_value(item.get(field_key, "")) or "-"
            for field_key, _ in item_fields
        ]
        row_number = header_row_index + index
        for col_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_number, column=col_index, value=value)
            cell.alignment = body_alignment
            cell.border = thin_border

    sheet.freeze_panes = "A5"
    last_row = max(header_row_index + len(data_rows), header_row_index)
    sheet.auto_filter.ref = f"A{header_row_index}:{last_col_letter}{last_row}"
    for col_index, (_, label) in enumerate(header_fields, start=1):
        width = 18
        if label in {"Lieferanschrift", "Store Name", "Store Address", "Mail To"}:
            width = 32
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    item_start_col = len(header_fields) + 1
    for offset, (_, label) in enumerate(item_fields, start=0):
        width = 18 if label != "Furncloud ID" else 20
        sheet.column_dimensions[get_column_letter(item_start_col + offset)].width = width

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _as_csv_text(orders: list[dict[str, Any]]) -> str:
    fieldnames = [
        "received_at",
        "status",
        "validation_status",
        "validation_summary",
        "validation_checked_at",
        "validation_provider",
        "validation_model",
        "validation_stale_reason",
        "validation_issues",
        "ticket_number",
        "kom_nr",
        "kom_name",
        "message_id",
        "kundennummer",
        "store_name",
        "store_address",
        "delivery_week",
        "liefertermin",
        "wunschtermin",
        "item_count",
        "warnings_count",
        "errors_count",
        "reply_needed",
        "human_review_needed",
        "post_case",
        "warnings",
        "errors",
        "file_name",
        "order_id",
    ]

    payload_map = _preload_order_export_payloads(orders)
    parsed_orders = [_load_order_export_data(order, payload_map=payload_map) for order in orders]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for parsed in parsed_orders:
        header = parsed.get("header", {}) if isinstance(parsed.get("header"), dict) else {}
        writer.writerow(
            {
                "received_at": parsed.get("received_at", ""),
                "status": _normalize_status(parsed.get("status")),
                "validation_status": parsed.get("validation_status", "not_run"),
                "validation_summary": parsed.get("validation_summary", ""),
                "validation_checked_at": parsed.get("validation_checked_at", ""),
                "validation_provider": parsed.get("validation_provider", ""),
                "validation_model": parsed.get("validation_model", ""),
                "validation_stale_reason": parsed.get("validation_stale_reason", ""),
                "validation_issues": json.dumps(parsed.get("validation_issues", []), ensure_ascii=False),
                "ticket_number": _export_entry_value(header.get("ticket_number", "")),
                "kom_nr": _export_entry_value(header.get("kom_nr", "")),
                "kom_name": _export_entry_value(header.get("kom_name", "")),
                "message_id": parsed.get("message_id", ""),
                "kundennummer": _export_entry_value(header.get("kundennummer", "")),
                "store_name": _export_entry_value(header.get("store_name", "")),
                "store_address": _export_entry_value(header.get("store_address", "")),
                "delivery_week": _export_entry_value(header.get("delivery_week", "")),
                "liefertermin": _export_entry_value(header.get("liefertermin", "")),
                "wunschtermin": _export_entry_value(header.get("wunschtermin", "")),
                "item_count": parsed.get("item_count", 0),
                "warnings_count": parsed.get("warnings_count", 0),
                "errors_count": parsed.get("errors_count", 0),
                "reply_needed": bool(parsed.get("reply_needed")),
                "human_review_needed": bool(parsed.get("human_review_needed")),
                "post_case": bool(parsed.get("post_case")),
                "warnings": " | ".join([str(item) for item in parsed.get("warnings", [])]),
                "errors": " | ".join([str(item) for item in parsed.get("errors", [])]),
                "file_name": parsed.get("file_name", ""),
                "order_id": parsed.get("order_id", ""),
            }
        )
    return output.getvalue()


def _data_export_columns(table_name: str, rows: list[dict[str, Any]]) -> list[str]:
    if rows:
        return [str(column) for column in rows[0].keys()]

    schema_rows = fetch_all(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = %s
        ORDER BY ordinal_position
        """,
        (table_name,),
    )
    return [str(row.get("column_name") or "").strip() for row in schema_rows if str(row.get("column_name") or "").strip()]


def _load_data_export_rows(table_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    if table_name not in ALLOWED_DATA_EXPORT_TABLES:
        raise ValueError(f"Unsupported export table: {table_name}")
    rows = fetch_all(f'SELECT * FROM "{table_name}"')
    export_map = IMPORT_COLUMN_MAP.get(table_name)
    if export_map:
        columns = list(export_map.keys())
        transformed_rows = [{header: row.get(db_column) for header, db_column in export_map.items()} for row in rows]
        return columns, transformed_rows
    columns = _data_export_columns(table_name, rows)
    return columns, rows


def _export_table_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return value.hex()
    return value


def _as_table_xlsx_bytes(*, table_name: str, columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = table_name[:31] or "Export"

    for column_index, column_name in enumerate(columns, start=1):
        header_cell = sheet.cell(row=1, column=column_index, value=column_name)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="left", vertical="bottom")

    for row_index, row in enumerate(rows, start=2):
        for column_index, column_name in enumerate(columns, start=1):
            value = _export_table_cell_value(row.get(column_name))
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    sheet.freeze_panes = "A2"

    min_width = 10
    max_width = 60
    padding = 2
    for column_index, column_name in enumerate(columns, start=1):
        width = len(str(column_name))
        for row in rows:
            text = "" if row.get(column_name) is None else str(_export_table_cell_value(row.get(column_name)))
            width = max(width, len(text))
        sheet.column_dimensions[get_column_letter(column_index)].width = max(min_width, min(max_width, width + padding))

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@app.before_request
def _api_auth_guard():
    if not request.path.startswith("/api/"):
        return None
    if request.path in {
        "/api/auth/login",
        "/api/auth/logout",
        "/api/auth/me",
        "/api/auth/check",
    }:
        return None
    return require_auth(request)


@app.after_request
def _api_cors_headers(response: Response):
    if not request.path.startswith("/api/"):
        return response

    origin = request.headers.get("Origin")
    if _is_origin_allowed(origin):
        if origin:
            response.headers["Access-Control-Allow-Origin"] = str(origin)
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Vary"] = _append_vary(response.headers.get("Vary"), "Origin")
        else:
            response.headers["Access-Control-Allow-Origin"] = "*"

    response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
    response.headers["Access-Control-Max-Age"] = "86400"
    return response


@app.errorhandler(HTTPException)
def _http_error_handler(error: HTTPException):
    if not request.path.startswith("/api/"):
        return error
    code_map = {
        400: "bad_request",
        401: "unauthorized",
        403: "forbidden",
        404: "not_found",
        405: "method_not_allowed",
    }
    status_code = error.code or 500
    code = code_map.get(status_code, "http_error")
    return _api_error(status_code, code, error.description or error.name)


@app.errorhandler(500)
def _internal_error_handler(error):
    if not request.path.startswith("/api/"):
        return error
    return _api_error(500, "internal_error", "Unexpected server error")


@app.route("/api", methods=["OPTIONS"])
@app.route("/api/<path:any_path>", methods=["OPTIONS"])
def api_options(any_path: str | None = None):  # noqa: ARG001
    return ("", 204)


@app.route("/api/auth/check")
def api_auth_check():
    session_id = request.cookies.get(session_cookie_name(), "")
    user = get_session_user(session_id)
    if not user:
        return _api_error(401, "unauthorized", "Authentication required")
    return ("", 204)


@app.route("/api/auth/me")
def api_auth_me():
    session_id = request.cookies.get(session_cookie_name(), "")
    user = get_session_user(session_id)
    if not user:
        return _api_error(401, "unauthorized", "Authentication required")
    return jsonify(
        {
            "user": {
                "id": user["id"],
                "username": user["username"],
                "role": user["role"],
                "client_branches": user.get("client_branches", []),
            }
        }
    )


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    payload = request.get_json(silent=True) or {}
    username = _clean_form_value(payload.get("username"))
    password = _clean_form_value(payload.get("password"))
    if not username or not password:
        return _api_error(400, "bad_request", "Username and password are required")

    user = authenticate_user(username, password)
    if not user:
        return _api_error(401, "unauthorized", "Invalid username or password")

    session_id = create_session(
        user["id"],
        request.remote_addr,
        request.headers.get("User-Agent"),
    )
    response = jsonify({"user": user})
    response.set_cookie(session_cookie_name(), session_id, **session_cookie_options())
    return response


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session_id = request.cookies.get(session_cookie_name(), "")
    if session_id:
        revoke_session(session_id)
    response = ("", 204)
    flask_response = app.make_response(response)
    flask_response.set_cookie(session_cookie_name(), "", max_age=0, path="/")
    return flask_response


@app.route("/api/settings/delivery-preparation", methods=["GET", "PUT", "POST"])
def api_delivery_preparation_settings():
    admin_error = _require_admin()
    if admin_error:
        return admin_error

    if request.method == "GET":
        try:
            return jsonify(get_delivery_preparation_settings())
        except Exception as exc:  # noqa: BLE001
            return _api_error(500, "db_error", f"Failed to load delivery preparation settings: {exc}")

    payload = request.get_json(silent=True)
    if payload is None:
        return _api_error(400, "bad_request", "Request body must be valid JSON")

    try:
        settings = replace_delivery_preparation_settings(payload)
    except ValueError as exc:
        return _api_error(400, "bad_request", str(exc))
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to save delivery preparation settings: {exc}")
    return jsonify(settings)


@app.route("/api/users", methods=["GET", "POST"])
def api_users():
    admin_error = _require_admin()
    if admin_error:
        return admin_error

    if request.method == "GET":
        rows = fetch_all(
            """
            SELECT id, username, email, role, is_active, created_at, updated_at, last_login_at
            FROM users
            ORDER BY created_at DESC
            """
        )
        users = [_serialize_user_record(row) for row in rows]
        return jsonify({"users": users})

    payload = request.get_json(silent=True) or {}
    username = _clean_form_value(payload.get("username"))
    password = _clean_form_value(payload.get("password"))
    email = _clean_form_value(payload.get("email")) or None
    role = _clean_form_value(payload.get("role")) or "user"
    is_active = payload.get("is_active")
    is_active = True if is_active is None else bool(is_active)
    _has_client_branches, raw_client_branches = _extract_client_branches_payload(payload)
    client_branches, branch_error = _parse_client_branches_input(raw_client_branches)
    if branch_error is not None:
        return branch_error

    if not username or not password:
        return _api_error(400, "bad_request", "Username and password are required")
    if role not in {"user", "admin"}:
        return _api_error(400, "bad_request", "Role must be 'user' or 'admin'")
    if role != "admin" and not client_branches:
        return _api_error(400, "bad_request", "client_branches is required for role 'user'")

    existing = fetch_one(
        "SELECT id FROM users WHERE lower(username) = lower(%s)",
        (username,),
    )
    if existing:
        return _api_error(400, "conflict", "Username already exists")

    if email:
        existing_email = fetch_one(
            "SELECT id FROM users WHERE lower(email) = lower(%s)",
            (email,),
        )
        if existing_email:
            return _api_error(400, "conflict", "Email already exists")

    now = datetime.now().astimezone()
    new_id = str(uuid.uuid4())
    execute(
        """
        INSERT INTO users (id, username, password_hash, email, role, is_active, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            new_id,
            username,
            hash_password(password),
            email,
            role,
            is_active,
            now,
            now,
        ),
    )
    if role == "admin":
        _replace_user_client_scopes(new_id, set(), now=now)
    else:
        _replace_user_client_scopes(new_id, client_branches or set(), now=now)
    return (
        jsonify(
            {
                "user": {
                    "id": new_id,
                    "username": username,
                    "email": email,
                    "role": role,
                    "is_active": is_active,
                    "created_at": now.isoformat(),
                    "client_branches": [] if role == "admin" else sorted(client_branches or set()),
                }
            }
        ),
        201,
    )


@app.route("/api/users/<user_id>", methods=["PATCH"])
def api_user_update(user_id: str):
    admin_error = _require_admin()
    if admin_error:
        return admin_error

    payload = request.get_json(silent=True) or {}
    username = _clean_form_value(payload.get("username")) or None
    password = _clean_form_value(payload.get("password")) or None
    email = _clean_form_value(payload.get("email")) or None
    role = _clean_form_value(payload.get("role")) or None
    is_active = payload.get("is_active")
    is_active = None if is_active is None else bool(is_active)
    has_client_branches, raw_client_branches = _extract_client_branches_payload(payload)
    client_branches, branch_error = _parse_client_branches_input(raw_client_branches)
    if has_client_branches and branch_error is not None:
        return branch_error

    existing = fetch_one("SELECT id, username, role FROM users WHERE id = %s", (user_id,))
    if not existing:
        return _api_error(404, "not_found", "User not found")

    current_user = getattr(g, "user", {}) or {}
    if current_user.get("id") == user_id:
        if is_active is False:
            return _api_error(400, "bad_request", "You cannot deactivate your own account")
        if role and role != "admin":
            return _api_error(400, "bad_request", "You cannot remove your own admin role")

    if role and role not in {"user", "admin"}:
        return _api_error(400, "bad_request", "Role must be 'user' or 'admin'")

    effective_role = role or str(existing.get("role") or "user")
    existing_client_branches = set(_fetch_user_client_branches(user_id))
    if effective_role == "admin":
        effective_client_branches: set[str] = set()
    elif has_client_branches:
        if not client_branches:
            return _api_error(400, "bad_request", "client_branches is required for role 'user'")
        effective_client_branches = set(client_branches)
    else:
        if not existing_client_branches:
            return _api_error(400, "bad_request", "client_branches is required for role 'user'")
        effective_client_branches = existing_client_branches

    if username:
        conflict = fetch_one(
            "SELECT id FROM users WHERE lower(username) = lower(%s) AND id <> %s",
            (username, user_id),
        )
        if conflict:
            return _api_error(400, "conflict", "Username already exists")

    if email:
        conflict = fetch_one(
            "SELECT id FROM users WHERE lower(email) = lower(%s) AND id <> %s",
            (email, user_id),
        )
        if conflict:
            return _api_error(400, "conflict", "Email already exists")

    fields = []
    values: list[Any] = []
    if username is not None:
        fields.append("username = %s")
        values.append(username)
    if email is not None:
        fields.append("email = %s")
        values.append(email)
    if role is not None:
        fields.append("role = %s")
        values.append(role)
    if is_active is not None:
        fields.append("is_active = %s")
        values.append(is_active)
    if password:
        fields.append("password_hash = %s")
        values.append(hash_password(password))

    now = datetime.now().astimezone()
    scopes_need_update = effective_client_branches != existing_client_branches
    if not fields and not scopes_need_update:
        return _api_error(400, "bad_request", "No changes provided")

    if fields:
        fields.append("updated_at = %s")
        values.append(now)
        values.append(user_id)
        execute(f"UPDATE users SET {', '.join(fields)} WHERE id = %s", values)
    elif scopes_need_update:
        execute("UPDATE users SET updated_at = %s WHERE id = %s", (now, user_id))

    _replace_user_client_scopes(user_id, effective_client_branches, now=now)
    updated_row = fetch_one(
        """
        SELECT id, username, email, role, is_active, created_at, updated_at, last_login_at
        FROM users
        WHERE id = %s
        """,
        (user_id,),
    )
    if not updated_row:
        return _api_error(404, "not_found", "User not found")
    return jsonify({"user": _serialize_user_record(updated_row)})


@app.route("/api/review/tasks", methods=["GET"])
def api_review_tasks():
    user = getattr(g, "user", {}) or {}
    scope = _order_access_scope(user, include_assignment=True)
    is_admin = scope["is_admin"]
    raw_states = (request.args.get("state") or "").strip()
    states: set[str] | None = None
    if raw_states:
        states = {item.strip().lower() for item in raw_states.split(",") if item.strip()}

    mine_only = _parse_bool_query(request.args.get("mine")) is True
    assigned_user_id: str | None = None
    include_unassigned = not mine_only
    allowed_client_branches = scope.get("allowed_client_branches")

    if not is_admin:
        assigned_user_id = scope.get("assigned_user_id")
        include_unassigned = False
    elif mine_only:
        assigned_user_id = str(user.get("id") or "")
    else:
        requested_user_id = (request.args.get("assigned_user_id") or "").strip()
        if requested_user_id:
            assigned_user_id = requested_user_id

    try:
        tasks = order_store.list_review_tasks(
            states=states,
            assigned_user_id=assigned_user_id,
            include_unassigned=include_unassigned,
            allowed_client_branches=allowed_client_branches,
        )
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to load review tasks: {exc}")
    return jsonify({"tasks": tasks})


@app.route("/api/review/tasks/<task_id>/claim", methods=["POST"])
def api_review_task_claim(task_id: str):
    try:
        uuid.UUID(task_id)
    except ValueError:
        return _api_error(404, "not_found", "Review task not found")
    user = getattr(g, "user", {}) or {}
    body = request.get_json(silent=True) or {}
    lease_seconds = body.get("lease_seconds", 300)
    try:
        task = order_store.claim_task(
            task_id=task_id,
            user_id=str(user.get("id") or ""),
            lease_seconds=int(lease_seconds),
        )
    except order_store.OrderStoreError as exc:
        return _order_store_error_response(exc)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to claim task: {exc}")
    return jsonify({"task": task})


@app.route("/api/review/tasks/<task_id>/heartbeat", methods=["POST"])
def api_review_task_heartbeat(task_id: str):
    try:
        uuid.UUID(task_id)
    except ValueError:
        return _api_error(404, "not_found", "Review task not found")
    user = getattr(g, "user", {}) or {}
    body = request.get_json(silent=True) or {}
    lease_seconds = body.get("lease_seconds", 300)
    try:
        task = order_store.heartbeat_task(
            task_id=task_id,
            user_id=str(user.get("id") or ""),
            lease_seconds=int(lease_seconds),
        )
    except order_store.OrderStoreError as exc:
        return _order_store_error_response(exc)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to heartbeat task: {exc}")
    return jsonify({"task": task})


@app.route("/api/review/tasks/<task_id>/resolve", methods=["POST"])
def api_review_task_resolve(task_id: str):
    try:
        uuid.UUID(task_id)
    except ValueError:
        return _api_error(404, "not_found", "Review task not found")
    user = getattr(g, "user", {}) or {}
    body = request.get_json(silent=True) or {}
    try:
        task = order_store.resolve_task(
            task_id=task_id,
            user_id=str(user.get("id") or ""),
            is_admin=user.get("role") == "admin",
            outcome=str(body.get("outcome") or "resolved"),
            note=str(body.get("note") or ""),
            force=bool(body.get("force")),
        )
    except order_store.OrderStoreError as exc:
        return _order_store_error_response(exc)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to resolve task: {exc}")
    return jsonify({"task": task})


@app.route("/api/overview")
def api_overview():
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    now = datetime.now().astimezone()
    overview_range, error_response = _parse_overview_range(now)
    if error_response is not None:
        return error_response

    try:
        overview = order_store.query_overview_snapshot(
            range_start=overview_range["start"],
            range_end=overview_range["end"],
            chart_start=overview_range["chart_start"],
            chart_end=overview_range["chart_end"],
            bucket_granularity=overview_range["bucket_granularity"],
            local_timezone=_postgres_timezone_name(now.tzinfo),
            assigned_user_id=scope.get("assigned_user_id"),
            allowed_client_branches=scope.get("allowed_client_branches"),
        )
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to load overview metrics: {exc}")

    summary = overview.get("summary", {})
    status_by_day_rows = overview.get("status_by_day", [])
    orders_by_client_hour_payload = overview.get("orders_by_client_hour", {}) or {}
    orders_by_client_hour_days = orders_by_client_hour_payload.get("days", []) or []

    status_by_day = []
    for row in status_by_day_rows:
        bucket_start = row.get("bucket_start")
        if isinstance(bucket_start, datetime):
            bucket_point = bucket_start.astimezone()
        elif isinstance(bucket_start, date):
            bucket_point = datetime.combine(bucket_start, datetime.min.time(), tzinfo=overview_range["start"].tzinfo)
        else:
            bucket_point = overview_range["start"]
        if overview_range["bucket_granularity"] == "month":
            label = bucket_point.strftime("%b %Y")
            iso_value = bucket_point.date().isoformat()
        else:
            label = bucket_point.strftime("%b %d")
            iso_value = bucket_point.date().isoformat()
        status_by_day.append(
            {
                "date": iso_value,
                "label": label,
                "ok": int(row.get("ok") or 0),
                "waiting_for_reply": int(row.get("waiting_for_reply") or 0),
                "human_in_the_loop": int(row.get("human_in_the_loop") or 0),
                "post": int(row.get("post") or 0),
                "unknown": int(row.get("unknown") or 0),
                "failed": int(row.get("failed") or 0),
                "updated_after_reply": int(row.get("updated_after_reply") or 0),
                "total": int(row.get("total") or 0),
            }
        )

    orders_by_client_hour = []
    for day in orders_by_client_hour_days:
        day_date = str(day.get("date") or "")
        try:
            day_point = datetime.fromisoformat(day_date)
            day_label = day_point.strftime("%b %d")
        except ValueError:
            day_label = day_date
        hours = []
        for hour_entry in day.get("hours", []) or []:
            hour_value = int(hour_entry.get("hour") or 0)
            hours.append(
                {
                    "hour": hour_value,
                    "label": f"{hour_value:02d}:00",
                    "total": int(hour_entry.get("total") or 0),
                    "clients": [
                        {
                            "id": str(client.get("id") or ""),
                            "count": int(client.get("count") or 0),
                        }
                        for client in (hour_entry.get("clients") or [])
                    ],
                }
            )
        orders_by_client_hour.append(
            {
                "date": day_date,
                "label": day_label,
                "total": int(day.get("total") or 0),
                "hours": hours,
            }
        )

    return jsonify(
        {
            "generated_at": now.isoformat(),
            "range": {
                "preset": overview_range["preset"],
                "month": overview_range.get("month"),
                "year": overview_range.get("year"),
                "start": overview_range["start"].isoformat(),
                "end": overview_range["end"].isoformat(),
                "chart_start": overview_range["chart_range_start"].isoformat(),
                "chart_end": overview_range["chart_range_end"].isoformat(),
                "bucket_granularity": overview_range["bucket_granularity"],
            },
            "summary": _overview_status_summary_from_counts(
                total=int(summary.get("period_total") or 0),
                ok=int(summary.get("period_ok") or 0),
                waiting_for_reply=int(summary.get("period_waiting_for_reply") or 0),
                human_in_the_loop=int(summary.get("period_human_in_the_loop") or 0),
                post=int(summary.get("period_post") or 0),
                unknown=int(summary.get("period_unknown") or 0),
                failed=int(summary.get("period_failed") or 0),
                updated_after_reply=int(summary.get("period_updated_after_reply") or 0),
            ),
            "status_by_day": status_by_day,
            "orders_by_client_hour": {
                "clients": [
                    {
                        "id": str(client.get("id") or ""),
                        "label": str(client.get("label") or client.get("id") or ""),
                    }
                    for client in (orders_by_client_hour_payload.get("clients") or [])
                ],
                "days": orders_by_client_hour,
            },
        }
    )


@app.route("/api/clients/counts")
def api_clients_counts():
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    try:
        counts = order_store.list_client_branch_counts(
            assigned_user_id=scope.get("assigned_user_id"),
            allowed_client_branches=scope.get("allowed_client_branches"),
        )
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to load client counts: {exc}")
    total = sum(int(value or 0) for value in counts.values())
    return jsonify({"counts": counts, "total": total})


@app.route("/api/orders")
def api_orders():
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    result, error = _query_orders(allow_default_pagination=True, access_scope=scope)
    if error is not None:
        return error

    return jsonify(
        {
            "orders": result["orders_serialized"],
            "pagination": result["pagination"],
            "counts": result["counts"],
        }
    )


@app.route("/api/orders.csv")
def api_orders_csv():
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    result, error = _query_orders(allow_default_pagination=False, access_scope=scope)
    if error is not None:
        return error

    try:
        csv_text = _as_csv_text(result["orders"])
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to build CSV export: {exc}")
    response = Response(csv_text, mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=orders.csv"
    return response


@app.route("/api/orders.xlsx")
def api_orders_xlsx():
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    result, error = _query_orders(allow_default_pagination=False, access_scope=scope)
    if error is not None:
        return error

    raw_title = (request.args.get("title") or "Orders").strip()
    initials = (request.args.get("initials") or "").strip()
    safe_title = re.sub(r"[^A-Za-z0-9_-]+", "_", raw_title).strip("_") or "Orders"
    date_stamp = datetime.now().astimezone().strftime("%d_%m_%y")
    filename_parts = [safe_title, date_stamp]
    if initials:
        filename_parts.append(initials)
    filename = "_".join(filename_parts) + ".xlsx"

    try:
        xlsx_bytes = _as_orders_xlsx_bytes(
            result["orders"],
            title=raw_title,
            initials=initials,
        )
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to build XLSX export: {exc}")
    response = Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/api/orders/export-xml.zip")
def api_orders_export_xml_zip():
    import io, zipfile
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    result, error = _query_orders(allow_default_pagination=False, access_scope=scope)
    if error is not None:
        return error

    orders = result["orders"]
    payload_map = _preload_order_export_payloads(orders)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for order in orders:
            order_id = str(order.get("id") or "").strip()
            entry = payload_map.get(order_id)
            if not entry or entry.get("parse_error"):
                continue
            data = entry["data"]
            try:
                documents = xml_exporter.render_xml_documents(data, "", config, Path("."))
            except Exception:
                continue
            for doc in documents:
                zf.writestr(doc.filename, doc.content)

    date_stamp = datetime.now().astimezone().strftime("%Y-%m-%d")
    filename = f"orders_xml_{date_stamp}.zip"
    zip_buffer.seek(0)
    response = Response(zip_buffer.read(), mimetype="application/zip")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/api/data-export/<table_name>.xlsx")
def api_data_export_table_xlsx(table_name: str):
    normalized_table_name = str(table_name or "").strip()
    target_table_name = DATA_EXPORT_TABLE_ALIASES.get(normalized_table_name, normalized_table_name)
    if target_table_name not in ALLOWED_DATA_EXPORT_TABLES:
        return _api_error(
            400,
            "invalid_table",
            f"Unsupported export table '{normalized_table_name}'",
        )

    try:
        columns, rows = _load_data_export_rows(target_table_name)
        xlsx_bytes = _as_table_xlsx_bytes(
            table_name=target_table_name,
            columns=columns,
            rows=rows,
        )
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to build XLSX export: {exc}")

    date_stamp = datetime.now().astimezone().strftime("%Y-%m-%d")
    filename = f"{target_table_name}_{date_stamp}.xlsx"
    response = Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _import_table_from_xlsx(table_name: str, file_stream: Any) -> int:
    from openpyxl import load_workbook

    col_map = IMPORT_COLUMN_MAP[table_name]
    wb = load_workbook(filename=file_stream, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return 0

    # Map positional index -> db column name (skip unrecognized headers)
    index_to_col: dict[int, str] = {}
    matched_headers: set[str] = set()
    for idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        header = str(cell_val).strip()
        if header in col_map:
            index_to_col[idx] = col_map[header]
            matched_headers.add(header)

    missing_headers = [header for header in col_map if header not in matched_headers]
    if missing_headers:
        missing_str = ", ".join(missing_headers)
        raise ValueError(f"Missing required columns: {missing_str}")

    data_rows: list[dict[str, Any]] = []
    for row in rows_iter:
        record: dict[str, Any] = {}
        for idx, db_col in index_to_col.items():
            cell_val = row[idx] if idx < len(row) else None
            if cell_val is None:
                record[db_col] = None
            elif isinstance(cell_val, float) and cell_val.is_integer():
                record[db_col] = str(int(cell_val))
            else:
                record[db_col] = str(cell_val).strip()
        if any(v is not None for v in record.values()):
            data_rows.append(record)

    db_cols = list(col_map.values())
    cols_sql = ", ".join(f'"{c}"' for c in db_cols)
    placeholders = ", ".join(["%s"] * len(db_cols))

    with transaction() as conn:
        with conn.cursor() as cursor:
            cursor.execute(f'DELETE FROM "{table_name}"')
            if data_rows:
                insert_rows = [[record.get(column_name) for column_name in db_cols] for record in data_rows]
                cursor.executemany(
                    f'INSERT INTO "{table_name}" ({cols_sql}) VALUES ({placeholders})',
                    insert_rows,
                )
    return len(data_rows)


@app.route("/api/data-import/<table_name>", methods=["POST"])
def api_data_import_table(table_name: str):
    normalized = str(table_name or "").strip()
    if normalized not in ALLOWED_DATA_IMPORT_TABLES:
        return _api_error(400, "invalid_table", f"Unsupported import table '{normalized}'")
    if "file" not in request.files:
        return _api_error(400, "missing_file", "No file uploaded")
    file = request.files["file"]
    try:
        count = _import_table_from_xlsx(normalized, file.stream)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "import_error", f"Import failed: {exc}")
    return jsonify({"imported": count})


@app.route("/api/orders/<order_id>", methods=["GET", "PATCH", "DELETE"])
def api_order_detail(order_id: str):
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    order, load_error = _load_order(
        order_id,
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
    )
    if load_error is not None:
        return load_error

    current_user = getattr(g, "user", {}) or {}
    current_user_id = str(current_user.get("id") or "")
    is_admin = current_user.get("role") == "admin"
    can_edit, _edit_reason = (False, "Order is not editable")
    if current_user_id:
        can_edit, _edit_reason = order_store.is_order_editable_for_detail(
            order=order,
            user_id=current_user_id,
            is_admin=is_admin,
        )
    order["is_editable"] = bool(can_edit)
    order["editability_reason"] = "" if can_edit else str(_edit_reason or "Order is not editable")

    if request.method == "GET":
        return jsonify(_order_api_payload(order))

    if request.method == "DELETE":
        try:
            deleted = order_store.soft_delete_order(order_id=order["safe_id"], actor_user_id=current_user_id or None)
        except Exception as exc:  # noqa: BLE001
            return _api_error(500, "db_error", f"Failed to delete order: {exc}")
        if not deleted:
            return _api_error(404, "not_found", "Order not found")
        _invalidate_order_index_cache()
        return jsonify({"deleted": True, "order_id": order["safe_id"]})

    if not can_edit:
        return _api_error(403, "forbidden", _edit_reason or "Order is not editable")

    body = request.get_json(silent=True)
    if not isinstance(body, dict):
        return _api_error(400, "invalid_body", "PATCH body must be a JSON object")

    header_updates = body.get("header", {})
    item_updates = body.get("items", {})
    deleted_item_indexes = body.get("deleted_item_indexes", [])
    new_items = body.get("new_items", [])
    if not isinstance(header_updates, dict):
        return _api_error(400, "invalid_body", "'header' must be an object")
    if not isinstance(item_updates, dict):
        return _api_error(400, "invalid_body", "'items' must be an object keyed by item index")
    if deleted_item_indexes is None:
        deleted_item_indexes = []
    if not isinstance(deleted_item_indexes, list):
        return _api_error(400, "invalid_body", "'deleted_item_indexes' must be an array")
    if new_items is None:
        new_items = []
    if not isinstance(new_items, list):
        return _api_error(400, "invalid_body", "'new_items' must be an array")

    for field, value in header_updates.items():
        if field not in EDITABLE_HEADER_FIELDS:
            return _api_error(400, "invalid_field", f"Header field '{field}' is not editable")
        _set_manual_entry(order["header"], field, _clean_form_value(str(value) if value is not None else ""))

    deleted_index_set: set[int] = set()
    item_count = len(order["items"])
    for raw_deleted_index in deleted_item_indexes:
        if isinstance(raw_deleted_index, bool):
            return _api_error(400, "invalid_body", f"Invalid deleted item index '{raw_deleted_index}'")
        try:
            deleted_index = int(raw_deleted_index)
        except (TypeError, ValueError):
            return _api_error(400, "invalid_body", f"Invalid deleted item index '{raw_deleted_index}'")
        if deleted_index < 0 or deleted_index >= item_count:
            return _api_error(400, "invalid_body", f"Deleted item index '{raw_deleted_index}' is out of range")
        deleted_index_set.add(deleted_index)

    parsed_item_updates: dict[int, dict[str, Any]] = {}
    for raw_index, fields in item_updates.items():
        if not isinstance(fields, dict):
            return _api_error(400, "invalid_body", f"Item patch for index '{raw_index}' must be an object")
        try:
            index = int(raw_index)
        except (TypeError, ValueError):
            return _api_error(400, "invalid_body", f"Invalid item index '{raw_index}'")
        if index < 0 or index >= item_count:
            return _api_error(400, "invalid_body", f"Item index '{raw_index}' is out of range")
        if index in deleted_index_set:
            return _api_error(400, "invalid_body", f"Item index '{raw_index}' cannot be patched and deleted")
        parsed_item_updates[index] = fields

    for index, fields in parsed_item_updates.items():
        item = order["items"][index]
        if not isinstance(item, dict):
            return _api_error(400, "invalid_body", f"Item '{index}' is not editable")

        for field, value in fields.items():
            if field not in EDITABLE_ITEM_FIELDS:
                return _api_error(400, "invalid_field", f"Item field '{field}' is not editable")
            _set_manual_entry(item, field, _clean_form_value(str(value) if value is not None else ""))
    if deleted_index_set:
        order["items"] = [item for idx, item in enumerate(order["items"]) if idx not in deleted_index_set]

    appended_new_items: list[dict[str, Any]] = []
    appended_item_refs: list[tuple[dict[str, Any], dict[str, str]]] = []
    for index, raw_new_item in enumerate(new_items):
        if not isinstance(raw_new_item, dict):
            return _api_error(400, "invalid_body", f"new_items[{index}] must be an object")

        invalid_fields = [field for field in raw_new_item.keys() if field not in EDITABLE_ITEM_FIELDS]
        if invalid_fields:
            return _api_error(400, "invalid_field", f"Item field '{invalid_fields[0]}' is not editable")

        normalized_values: dict[str, str] = {}
        for field in EDITABLE_ITEM_FIELDS:
            value = raw_new_item.get(field, "")
            normalized_values[field] = _clean_form_value(str(value) if value is not None else "")

        if not any(normalized_values.values()):
            continue

        appended_item: dict[str, Any] = {}
        for field, value in normalized_values.items():
            _set_manual_entry(appended_item, field, value)
        order["items"].append(appended_item)
        appended_item_refs.append((appended_item, normalized_values))

    for line_no, existing_item in enumerate(order["items"], start=1):
        if isinstance(existing_item, dict):
            existing_item["line_no"] = line_no

    for appended_item, normalized_values in appended_item_refs:
        appended_new_items.append(
            {
                "line_no": appended_item.get("line_no"),
                **normalized_values,
            }
        )

    order["data"]["header"] = order["header"]
    order["data"]["items"] = order["items"]
    refresh_missing_warnings(order["data"])

    revision_id: str | None = None
    try:
        persisted = order_store.save_manual_revision(
            order_id=order["safe_id"],
            payload=order["data"],
            actor_user_id=current_user_id,
            diff_json={
                "header": header_updates,
                "items": item_updates,
                "deleted_item_indexes": sorted(deleted_index_set),
                "new_items": appended_new_items,
            },
        )
        revision_id = persisted.get("revision_id")
    except order_store.OrderStoreError as exc:
        return _order_store_error_response(exc)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to save manual revision: {exc}")
    try:
        _mark_order_validation_stale(
            order["safe_id"],
            actor_user_id=current_user_id or None,
            reason="manual_edit",
        )
    except RuntimeError as exc:
        return _api_error(500, "db_error", str(exc))

    xml_regenerated = False
    xml_paths: list[str] = []
    try:
        xml_paths = [str(path) for path in xml_exporter.export_xmls(order["data"], order["safe_id"], config, OUTPUT_DIR)]
        xml_regenerated = True
    except Exception as exc:  # noqa: BLE001
        xml_regenerated = False
        try:
            order_store.record_order_event(
                order_id=order["safe_id"],
                revision_id=revision_id,
                event_type="xml_regeneration_failed",
                actor_user_id=current_user_id or None,
                event_data={"error": str(exc)},
            )
        except Exception as db_exc:  # noqa: BLE001
            return _api_error(500, "db_error", f"Failed to record XML regeneration failure metadata: {db_exc}")
    if xml_regenerated:
        try:
            order_store.register_order_files(
                order_id=order["safe_id"],
                revision_id=revision_id,
                file_type="xml",
                storage_paths=xml_paths,
            )
            order_store.record_order_event(
                order_id=order["safe_id"],
                revision_id=revision_id,
                event_type="xml_regenerated",
                actor_user_id=current_user_id or None,
                event_data={"files": xml_paths},
            )
        except Exception as exc:  # noqa: BLE001
            return _api_error(500, "db_error", f"Failed to record XML regeneration metadata: {exc}")

    _invalidate_order_index_cache()
    updated_order, updated_error = _load_order(
        order["safe_id"],
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
    )
    if updated_error is not None:
        return updated_error
    updated_can_edit, updated_reason = order_store.is_order_editable_for_detail(
        order=updated_order,
        user_id=current_user_id,
        is_admin=is_admin,
    )
    updated_order["is_editable"] = bool(updated_can_edit)
    updated_order["editability_reason"] = "" if updated_can_edit else str(updated_reason or "Order is not editable")

    payload = _order_api_payload(updated_order)
    payload["xml_regenerated"] = xml_regenerated
    return jsonify(payload)


@app.route("/api/orders/<order_id>/header.xlsx")
def api_export_order_header_xlsx(order_id: str):
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    order, load_error = _load_order(
        order_id,
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
    )
    if load_error is not None:
        return load_error

    try:
        xlsx_bytes = _as_order_header_xlsx_bytes(order)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "xlsx_export_failed", f"Failed to build header XLSX export: {exc}")

    base_name = str(order.get("safe_id") or order.get("id") or "order").strip() or "order"
    safe_base_name = re.sub(r"[^A-Za-z0-9_-]+", "_", base_name).strip("_") or "order"
    filename = f"{safe_base_name}_header_information.xlsx"
    response = Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@app.route("/api/orders/<order_id>/export-xml", methods=["POST"])
def api_export_order_xml(order_id: str):
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    order, load_error = _load_order(
        order_id,
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
    )
    if load_error is not None:
        return load_error
    if order["parse_error"]:
        return _api_error(400, "invalid_order", "Order payload could not be parsed")

    current_user = getattr(g, "user", {}) or {}
    actor_user_id = str(current_user.get("id") or "") or None

    try:
        xml_paths = [str(path) for path in xml_exporter.export_xmls(order["data"], order["safe_id"], config, OUTPUT_DIR)]
    except Exception:  # noqa: BLE001
        return _api_error(500, "xml_export_failed", "Failed to regenerate XML files")
    try:
        order_store.register_order_files(
            order_id=order["safe_id"],
            revision_id=None,
            file_type="xml",
            storage_paths=xml_paths,
        )
        order_store.record_order_event(
            order_id=order["safe_id"],
            event_type="xml_regenerated",
            actor_user_id=actor_user_id,
            event_data={"files": xml_paths},
        )
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to record XML export metadata: {exc}")
    try:
        _mark_order_validation_stale(
            order["safe_id"],
            actor_user_id=actor_user_id,
            reason="manual_xml_export",
        )
    except RuntimeError as exc:
        return _api_error(500, "db_error", str(exc))

    files = _resolve_xml_files(order["safe_id"], order["header"])
    return jsonify({"xml_files": files})


@app.route("/api/orders/<order_id>/validation/resolve", methods=["POST"])
def api_resolve_order_validation(order_id: str):
    scope = _order_access_scope(getattr(g, "user", {}) or {})
    order, load_error = _load_order(
        order_id,
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
    )
    if load_error is not None:
        return load_error

    body = request.get_json(silent=True)
    if body is None:
        body = {}
    if not isinstance(body, dict):
        return _api_error(400, "invalid_body", "POST body must be a JSON object")
    note = str(body.get("note") or "").strip()
    if not note:
        return _api_error(400, "invalid_body", "A resolution note is required")

    actor_user_id = str((getattr(g, "user", {}) or {}).get("id") or "") or None
    try:
        order_store.resolve_validation(
            order_id=order["safe_id"],
            actor_user_id=actor_user_id,
            note=note,
        )
    except order_store.OrderStoreError as exc:
        return _order_store_error_response(exc)
    except Exception as exc:  # noqa: BLE001
        return _api_error(500, "db_error", f"Failed to resolve validation: {exc}")

    _invalidate_order_index_cache()
    updated_order, updated_error = _load_order(
        order["safe_id"],
        assigned_user_id=scope.get("assigned_user_id"),
        allowed_client_branches=scope.get("allowed_client_branches"),
    )
    if updated_error is not None:
        return updated_error
    return jsonify(_order_api_payload(updated_order))


@app.route("/api/files/<filename>")
def api_download_file(filename: str):
    safe_filename = _safe_id(filename)
    if not safe_filename:
        return _api_error(404, "not_found", "File not found")

    extension = Path(safe_filename).suffix.lower()
    if extension not in ALLOWED_DOWNLOAD_EXTENSIONS:
        return _api_error(403, "forbidden", "File type is not allowed")

    full_path = OUTPUT_DIR / safe_filename
    if not full_path.exists() or not full_path.is_file():
        return _api_error(404, "not_found", "File not found")

    return send_from_directory(OUTPUT_DIR, safe_filename, as_attachment=True)

@app.route("/")
def index() -> str:
    try:
        orders = _get_order_index()
    except Exception:  # noqa: BLE001
        abort(500)

    date_scope = (request.args.get("date_scope") or "").lower().strip()
    if date_scope not in {"today", "all"}:
        date_scope = "today"

    for order in orders:
        parsed = _parse_received_at(order.get("received_at"))
        order["_received_at_sort"] = parsed

    orders_sorted = sorted(
        orders,
        key=lambda order: (
            order.get("_received_at_sort") is not None,
            order.get("_received_at_sort"),
        ),
        reverse=True,
    )

    if date_scope == "today":
        today = datetime.now().astimezone().date()
        scoped_orders = [
            order
            for order in orders_sorted
            if isinstance(order.get("_received_at_sort"), datetime)
            and order["_received_at_sort"].date() == today
        ]
    else:
        scoped_orders = orders_sorted

    counts = _status_counts(scoped_orders)

    status_filter = (request.args.get("status") or "").lower().strip()
    if status_filter and status_filter != "all":
        filtered_orders = [order for order in scoped_orders if order.get("status") == status_filter]
    else:
        status_filter = "all"
        filtered_orders = scoped_orders

    total_rows = len(filtered_orders)
    for idx, order in enumerate(filtered_orders, start=1):
        order["display_index"] = total_rows - idx + 1

    return render_template(
        "index.html",
        orders=filtered_orders,
        counts=counts,
        status_filter=status_filter,
        date_scope=date_scope,
        body_class="dashboard",
    )


@app.route("/excel-to-xml")
def excel_to_xml_page():
    return render_template("excel_to_xml.html", title="Excel → XML Generator")


@app.route("/excel-to-xml/generate", methods=["POST"])
def excel_to_xml_generate():
    import io
    import tempfile
    import zipfile

    from excel_xml_generator import generate_xmls_from_excel

    uploaded = request.files.get("excel_file")
    if not uploaded or not uploaded.filename:
        return render_template(
            "excel_to_xml.html",
            title="Excel → XML Generator",
            error="No file uploaded.",
        )

    filename_lower = uploaded.filename.lower()
    if not any(filename_lower.endswith(ext) for ext in (".xlsx", ".xlsb", ".xls")):
        return render_template(
            "excel_to_xml.html",
            title="Excel → XML Generator",
            error="Unsupported file type. Please upload .xlsx, .xlsb, or .xls.",
        )

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            xml_paths = generate_xmls_from_excel(uploaded.stream, Path(tmpdir))
            if not xml_paths:
                return render_template(
                    "excel_to_xml.html",
                    title="Excel → XML Generator",
                    error="No rows found in the Excel file.",
                )
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for p in xml_paths:
                    zf.write(p, arcname=p.name)
            zip_buffer.seek(0)
            return send_file(
                zip_buffer,
                mimetype="application/zip",
                as_attachment=True,
                download_name="orders_xml.zip",
            )
    except Exception as exc:  # noqa: BLE001
        return render_template(
            "excel_to_xml.html",
            title="Excel → XML Generator",
            error=f"Failed to generate XMLs: {exc}",
        )


@app.route("/download/<filename>")
def download_file(filename: str):
    safe_filename = _safe_id(filename)
    if not safe_filename:
        abort(404)
    extension = Path(safe_filename).suffix.lower()
    if extension not in ALLOWED_DOWNLOAD_EXTENSIONS:
        abort(403)
    return send_from_directory(OUTPUT_DIR, safe_filename, as_attachment=True)


@app.route("/order/<order_id>/export-xml", methods=["POST"])
def export_order_xml(order_id: str):
    order, load_error = _load_order(order_id)
    if load_error is not None:
        abort(_response_status_code(load_error, 500))
    if order["parse_error"]:
        abort(400)
    actor_user_id = str((getattr(g, "user", {}) or {}).get("id") or "") or None
    try:
        xml_paths = [str(path) for path in xml_exporter.export_xmls(order["data"], order["safe_id"], config, OUTPUT_DIR)]
        order_store.register_order_files(
            order_id=order["safe_id"],
            revision_id=None,
            file_type="xml",
            storage_paths=xml_paths,
        )
        order_store.record_order_event(
            order_id=order["safe_id"],
            event_type="xml_regenerated",
            actor_user_id=actor_user_id,
            event_data={"files": xml_paths},
        )
        _mark_order_validation_stale(
            order["safe_id"],
            actor_user_id=actor_user_id,
            reason="manual_xml_export",
        )
    except Exception:  # noqa: BLE001
        abort(500)
    _invalidate_order_index_cache()
    return redirect(url_for("order_detail", order_id=order["safe_id"], exported="1"))


@app.route("/order/<order_id>/delete", methods=["POST"])
def delete_order(order_id: str):
    order, load_error = _load_order(order_id)
    if load_error is not None:
        abort(_response_status_code(load_error, 500))
    try:
        deleted = order_store.soft_delete_order(order_id=order["safe_id"], actor_user_id=None)
    except Exception:  # noqa: BLE001
        abort(500)
    if not deleted:
        abort(404)
    _invalidate_order_index_cache()

    date_scope = (request.args.get("date_scope") or "").lower().strip()
    status_filter = (request.args.get("status") or "").lower().strip()
    return redirect(url_for("index", date_scope=date_scope or "today", status=status_filter or "all"))


@app.route("/order/<order_id>", methods=["GET", "POST"])
def order_detail(order_id: str) -> str:
    order, load_error = _load_order(order_id)
    if load_error is not None:
        abort(_response_status_code(load_error, 500))

    safe_id = order["safe_id"]
    data = order["data"]
    header = order["header"]
    items = order["items"]
    parse_error = str(order.get("parse_error") or "")
    human_review_needed = bool(order.get("human_review_needed"))
    reply_needed = bool(order.get("reply_needed"))
    post_case = bool(order.get("post_case"))

    if request.method == "POST":
        if parse_error or not human_review_needed:
            abort(403)

        for field in EDITABLE_HEADER_FIELDS:
            form_key = f"header_{field}"
            if form_key in request.form:
                _set_manual_entry(header, field, _clean_form_value(request.form.get(form_key)))

        for idx, item in enumerate(items):
            if not isinstance(item, dict):
                continue
            for field in EDITABLE_ITEM_FIELDS:
                form_key = f"item_{idx}_{field}"
                if form_key in request.form:
                    _set_manual_entry(item, field, _clean_form_value(request.form.get(form_key)))

        data["header"] = header
        data["items"] = items
        refresh_missing_warnings(data)
        actor_user_id = str((getattr(g, "user", {}) or {}).get("id") or "legacy_ui")
        revision_id: str | None = None
        try:
            persisted = order_store.save_manual_revision(
                order_id=safe_id,
                payload=data,
                actor_user_id=actor_user_id,
            )
            revision_id = persisted.get("revision_id")
            _mark_order_validation_stale(
                safe_id,
                actor_user_id=actor_user_id,
                reason="manual_edit",
            )
        except order_store.OrderStoreError as exc:
            abort(exc.status_code)
        except Exception:  # noqa: BLE001
            abort(500)

        xml_regenerated = False
        xml_paths: list[str] = []
        try:
            xml_paths = [str(path) for path in xml_exporter.export_xmls(data, safe_id, config, OUTPUT_DIR)]
            xml_regenerated = True
        except Exception:  # noqa: BLE001
            xml_regenerated = False
            try:
                order_store.record_order_event(
                    order_id=safe_id,
                    revision_id=revision_id,
                    event_type="xml_regeneration_failed",
                    actor_user_id=actor_user_id,
                    event_data={"error": "Failed to regenerate XML files from legacy detail route"},
                )
            except Exception:  # noqa: BLE001
                abort(500)
        if xml_regenerated:
            try:
                order_store.register_order_files(
                    order_id=safe_id,
                    revision_id=revision_id,
                    file_type="xml",
                    storage_paths=xml_paths,
                )
                order_store.record_order_event(
                    order_id=safe_id,
                    revision_id=revision_id,
                    event_type="xml_regenerated",
                    actor_user_id=actor_user_id,
                    event_data={"files": xml_paths},
                )
            except Exception:  # noqa: BLE001
                abort(500)

        _invalidate_order_index_cache()
        return redirect(
            url_for("order_detail", order_id=safe_id, saved="1", xml_regenerated="1" if xml_regenerated else "0")
        )

    header_rows = [
        {"field": "ticket_number", **_entry_dict(header.get("ticket_number"))},
        {"field": "kundennummer", **_entry_dict(header.get("kundennummer"))},
        {"field": "adressnummer", **_entry_dict(header.get("adressnummer"))},
        {"field": "tour", **_entry_dict(header.get("tour"))},
        {"field": "kom_nr", **_entry_dict(header.get("kom_nr"))},
        {"field": "kom_name", **_entry_dict(header.get("kom_name"))},
        {"field": "liefertermin", **_entry_dict(header.get("liefertermin"))},
        {"field": "wunschtermin", **_entry_dict(header.get("wunschtermin"))},
        {"field": "bestelldatum", **_entry_dict(header.get("bestelldatum"))},
        {"field": "lieferanschrift", **_entry_dict(header.get("lieferanschrift"))},
        {"field": "store_name", **_entry_dict(header.get("store_name"))},
        {"field": "store_address", **_entry_dict(header.get("store_address"))},
        {"field": "seller", **_entry_dict(header.get("seller"))},
        {"field": "delivery_week", **_entry_dict(header.get("delivery_week"))},
        {"field": "iln", **_entry_dict(header.get("iln"))},
        {"field": "human_review_needed", **_entry_dict(header.get("human_review_needed"))},
        {"field": "reply_needed", **_entry_dict(header.get("reply_needed"))},
        {"field": "post_case", **_entry_dict(header.get("post_case"))},
    ]

    item_rows = []
    for item in items:
        if not isinstance(item, dict):
            continue
        item_rows.append(
            {
                "line_no": item.get("line_no", ""),
                "artikelnummer": _entry_dict(item.get("artikelnummer")),
                "modellnummer": _entry_dict(item.get("modellnummer")),
                "menge": _entry_dict(item.get("menge")),
                "furncloud_id": _entry_dict(item.get("furncloud_id")),
            }
        )

    warnings = data.get("warnings", [])
    errors = data.get("errors", [])
    if not isinstance(warnings, list):
        warnings = [str(warnings)]
    warnings = [str(item) for item in warnings]
    if not isinstance(errors, list):
        errors = [str(errors)]
    errors = [str(item) for item in errors]

    reply_case = _reply_case_from_warnings(warnings)
    raw_json = json.dumps(data, ensure_ascii=False, indent=2)
    xml_files = _resolve_xml_files(safe_id, header)

    saved = (request.args.get("saved") or "") == "1"
    exported = (request.args.get("exported") or "") == "1"
    xml_regenerated = (request.args.get("xml_regenerated") or "") == "1"
    status = _normalize_status(data.get("status"))
    return render_template(
        "detail.html",
        order_id=safe_id,
        message_id=data.get("message_id") or safe_id,
        received_at=data.get("received_at") or "",
        status=status,
        status_label=_status_label(status),
        header_rows=header_rows,
        item_rows=item_rows,
        warnings=warnings,
        errors=errors,
        raw_json=raw_json,
        parse_error=parse_error,
        xml_files=xml_files,
        ab_files=[],
        is_editable=human_review_needed and not parse_error,
        reply_needed=reply_needed,
        post_case=post_case,
        reply_mailto=_reply_mailto(data.get("message_id") or safe_id, safe_id, reply_case) if reply_needed else "",
        editable_header_fields=EDITABLE_HEADER_FIELDS,
        editable_item_fields=EDITABLE_ITEM_FIELDS,
        saved=saved,
        exported=exported,
        xml_regenerated=xml_regenerated,
    )


if __name__ == "__main__":
    host = os.getenv("DASHBOARD_HOST", "127.0.0.1")
    port = int(os.getenv("DASHBOARD_PORT", "5000"))
    app.run(host=host, port=port, debug=False)
