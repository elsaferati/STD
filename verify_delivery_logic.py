import datetime
from pathlib import Path
from unittest.mock import patch

import openpyxl

import delivery_logic


WORKBOOK_PATH = Path(__file__).with_name("Lieferlogik_V2.xlsx")


def _load_workbook_reference() -> tuple[dict[str, str], dict[str, list[int]], dict[str, dict[int, int]]]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Kapa Base"]

    tour_to_code: dict[str, str] = {}
    for col in range(11, 17):
        tour = str(ws.cell(1, col).value or "").strip()
        code = str(ws.cell(2, col).value or "").strip()
        tour_to_code[tour] = code

    valid_weeks_by_code: dict[str, list[int]] = {}
    for col in range(2, 8):
        code = str(ws.cell(2, col).value or "").strip()
        valid_weeks: list[int] = []
        for row in range(3, 55):
            week = ws.cell(row, 1).value
            val = ws.cell(row, col).value
            if week is None or val is None:
                continue
            if float(val) > 0:
                valid_weeks.append(int(week))
        valid_weeks_by_code[code] = valid_weeks

    return tour_to_code, valid_weeks_by_code


class _FrozenDate(datetime.date):
    frozen_today = datetime.date(2026, 1, 5)

    @classmethod
    def today(cls) -> "_FrozenDate":
        return cls(
            cls.frozen_today.year,
            cls.frozen_today.month,
            cls.frozen_today.day,
        )


def _set_today(value: datetime.date) -> None:
    _FrozenDate.frozen_today = value


def _settings(default_prep_weeks: int = 2, ranges: list[dict[str, int]] | None = None) -> dict[str, object]:
    return {
        "default_prep_weeks": default_prep_weeks,
        "ranges": list(ranges or []),
    }


def test_schedule_tables_match_workbook_reference() -> None:
    expected_tour_to_code, expected_valid_weeks = _load_workbook_reference()

    assert delivery_logic.TOUR_TO_SCHEDULE_CODE == expected_tour_to_code
    assert {
        key: list(value) for key, value in delivery_logic.VALID_WEEKS_BY_CODE.items()
    } == expected_valid_weeks
    print("SUCCESS: delivery logic schedule constants match the workbook reference.")


def test_tour_validation_compatibility() -> None:
    for value in ("W1", "U2", "D1", "G2", "D2", "D3", "1.1", "1.2", "2.2", "1.3", "2.3", "3.3", "D2.1"):
        assert delivery_logic.is_tour_valid(value) is True
    for value in ("", "X9", "foo", "9.9"):
        assert delivery_logic.is_tour_valid(value) is False
    print("SUCCESS: delivery logic validates tours and aliases as expected.")


def test_delivery_week_without_requested_week() -> None:
    _set_today(datetime.date(2026, 1, 5))  # ISO week 2
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        result = delivery_logic.calculate_delivery_week("01.01.1999", "D2")
    assert result == "2026 Week - 05"
    print("SUCCESS: delivery week without requested week uses the default prep rule.")


def test_delivery_week_uses_custom_prep_range() -> None:
    _set_today(datetime.date(2026, 2, 16))  # ISO week 8
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(
            default_prep_weeks=2,
            ranges=[{"year_from": 2026, "week_from": 7, "year_to": 2026, "week_to": 17, "prep_weeks": 4}],
        ),
    ):
        result = delivery_logic.calculate_delivery_week("01.01.1999", "D2")
    assert result == "2026 Week - 14"
    print("SUCCESS: custom prep ranges override the default rule.")


def test_all_tours_pick_first_valid_week_on_or_after_candidate() -> None:
    _set_today(datetime.date(2026, 2, 16))  # ISO week 8
    expected = {
        "W1": "2026 Week - 10",
        "U2": "2026 Week - 11",
        "D1": "2026 Week - 10",
        "G2": "2026 Week - 10",
        "D2": "2026 Week - 11",
        "D3": "2026 Week - 12",
    }

    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        for tour, expected_value in expected.items():
            result = delivery_logic.calculate_delivery_week("01.01.1999", tour)
            assert result == expected_value, f"{tour}: expected {expected_value}, got {result}"
    print("SUCCESS: every tour uses the first valid service week on or after the candidate week.")


def test_delivery_week_requested_week_inside_window() -> None:
    _set_today(datetime.date(2026, 1, 5))  # ISO week 2
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        result = delivery_logic.calculate_delivery_week("01.01.1999", "G2", "KW06/2026")
    assert result == "2026 Week - 04"
    print("SUCCESS: requested-week logic still honors the earliest possible week.")


def test_delivery_week_requested_before_earliest_possible() -> None:
    _set_today(datetime.date(2026, 1, 5))  # ISO week 2
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=5),
    ):
        result = delivery_logic.calculate_delivery_week("01.01.1999", "D1", "KW05/2026")
    assert result == "2026 Week - 07"
    print("SUCCESS: requested week before earliest possible falls forward correctly.")


def test_delivery_week_falls_forward_when_window_has_no_valid_week() -> None:
    _set_today(datetime.date(2026, 11, 16))  # ISO week 47
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        result = delivery_logic.calculate_delivery_week("01.01.1999", "D1", "KW47/2026")
    assert result == "2026 Week - 49"
    print("SUCCESS: delivery week falls forward when the allowed window has no valid tour week.")


def test_requested_week_fallback_never_undercuts_earliest_possible_week() -> None:
    _set_today(datetime.date(2026, 3, 7))  # ISO week 10
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(
            default_prep_weeks=2,
            ranges=[{"year_from": 2026, "week_from": 10, "year_to": 2026, "week_to": 17, "prep_weeks": 7}],
        ),
    ):
        result = delivery_logic.calculate_delivery_week("17.02.2026", "G2", "11 KW 2026", client_name="segmuller")
    assert result == "2026 Week - 18"
    print("SUCCESS: requested-week fallback does not return a week before the earliest possible delivery.")


def test_braun_uses_shorter_early_offset() -> None:
    _set_today(datetime.date(2025, 12, 29))  # ISO week 1 of 2026
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        default_result = delivery_logic.calculate_delivery_week("01.01.1999", "U2", "KW10/2026", client_name="xxxlutz_default")
        braun_result = delivery_logic.calculate_delivery_week("01.01.1999", "U2", "KW10/2026", client_name="braun")
    assert default_result == "2026 Week - 05"
    assert braun_result == "2026 Week - 09"
    print("SUCCESS: Braun uses the shorter early-offset rule.")


def test_order_date_is_ignored() -> None:
    _set_today(datetime.date(2026, 1, 5))  # ISO week 2
    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        result_a = delivery_logic.calculate_delivery_week("01.01.2001", "D2")
        result_b = delivery_logic.calculate_delivery_week("31.12.2099", "D2")
    assert result_a == result_b == "2026 Week - 05"
    print("SUCCESS: order_date_str is ignored and today's week drives the calculation.")


def test_year_end_rollover_is_calendar_correct() -> None:
    scenarios = [
        (datetime.date(2026, 12, 21), "W1", "2027 Week - 01"),  # ISO week 52 + 2 prep weeks
        (datetime.date(2026, 12, 21), "D3", "2027 Week - 03"),
        (datetime.date(2026, 12, 28), "G2", "2027 Week - 02"),  # ISO week 53 falls through to default rule
    ]

    with patch("delivery_logic.datetime.date", _FrozenDate), patch(
        "delivery_logic._get_delivery_preparation_settings_cached",
        return_value=_settings(default_prep_weeks=2),
    ):
        for frozen_day, tour, expected in scenarios:
            _set_today(frozen_day)
            result = delivery_logic.calculate_delivery_week("01.01.1999", tour)
            assert result == expected
    print("SUCCESS: year-end rollover returns the correct ISO year.")


if __name__ == "__main__":
    test_schedule_tables_match_workbook_reference()
    test_tour_validation_compatibility()
    test_delivery_week_without_requested_week()
    test_delivery_week_uses_custom_prep_range()
    test_all_tours_pick_first_valid_week_on_or_after_candidate()
    test_delivery_week_requested_week_inside_window()
    test_delivery_week_requested_before_earliest_possible()
    test_delivery_week_falls_forward_when_window_has_no_valid_week()
    test_requested_week_fallback_never_undercuts_earliest_possible_week()
    test_braun_uses_shorter_early_offset()
    test_order_date_is_ignored()
    test_year_end_rollover_is_calendar_correct()
